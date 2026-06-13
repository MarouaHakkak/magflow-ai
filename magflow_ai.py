"""MagFlow AI v3 — Core Engine"""
import openpyxl, math
from dataclasses import dataclass, field
from typing import List, Dict, Optional

class DataLoader:
    def __init__(self, filepath):
        self.wb = openpyxl.load_workbook(filepath, data_only=True)
        self.fluids = self._load_fluids()
        self.electrodes = self._load_electrodes()
        self.liners = self._load_liners()
        self.vendor_models = self._load_models()
        self.drift_indicators = self._load_drift()
        self.capex_scores = self._load_capex()
        self.project_data = self._load_projects()
        self.asme = {150:{20:19.6,50:19.2,100:17.6,150:15.8,200:13.9,250:12.1,300:10.3},
            300:{20:51.1,50:49.7,100:45.9,150:42.1,200:38.2,250:34.5,300:30.7},
            600:{20:102.1,50:99.3,100:91.7,150:84.2,200:76.5,250:69.0,300:61.3},
            900:{20:153.1,50:149.0,100:137.6,150:126.2,200:114.7,250:103.4,300:91.9},
            1500:{20:255.1,50:248.3,100:229.5,150:210.4,200:191.2,250:172.4,300:153.3},
            2500:{20:425.2,50:414.0,100:382.5,150:351.0,200:319.4,250:287.8,300:256.3}}
        self.liner_dn = {'PFA':{'E+H':200,'Emerson':350,'Krohne':150},'PFA HT':{'E+H':200},'PFA+':{'Emerson':350},'PTFE':{'E+H':600,'Emerson':900,'Krohne':1200}}
    def _s(self, v): return str(v).strip() if v else ''
    def _load_fluids(self):
        ws = self.wb['Fluid-Material Matrix']; fluids = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            n = self._s(row[0])
            if n and len(n)>2 and not any(k in n for k in ['RÈGLES','SOURCE','NOTE','NOTES','1.','2.','3.','4.','5.','6.','7.','8.']):
                fluids.append({'name':n,'type':self._s(row[1]) if len(row)>1 else '','liner':self._s(row[4]) if len(row)>4 else '','electrode':self._s(row[5]) if len(row)>5 else '','penetrant':self._s(row[9]) if len(row)>9 else '','grounding':self._s(row[10]) if len(row)>10 else '','o_ring':self._s(row[11]) if len(row)>11 else '','flange_coat':self._s(row[12]) if len(row)>12 else '','pipe_mat':self._s(row[13]) if len(row)>13 else '','remarks':self._s(row[14]) if len(row)>14 else ''})
        return fluids
    def _load_electrodes(self):
        ws = self.wb['Electrode Materials']; items = []
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            n = self._s(row[0])
            if n: items.append({'name':n,'emerson':self._s(row[1]),'eh':self._s(row[2]),'krohne':self._s(row[3]),'availability':self._s(row[4]),'temp_acids':self._s(row[5]),'temp_water':self._s(row[6]),'abrasion':self._s(row[7]),'cost':self._s(row[8])})
        return items
    def _load_liners(self):
        ws = self.wb['Liner Materials']; items = []
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            n = self._s(row[0])
            if n: items.append({'name':n,'emerson':self._s(row[1]),'eh':self._s(row[2]),'krohne':self._s(row[3]),'availability':self._s(row[4]),'temp':self._s(row[5]),'abrasion':self._s(row[6]),'vacuum':self._s(row[7]),'cost':self._s(row[8])})
        return items
    def _load_models(self):
        ws = self.wb['Vendor Models']; models = []
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            v,m = self._s(row[0]), self._s(row[1])
            if v and m and v!='Vendor':
                d = {'vendor':v,'model':m,'type':self._s(row[2]),'eq_emerson':self._s(row[3]),'eq_eh':self._s(row[4]),'eq_krohne':self._s(row[5]),'electrodes':self._s(row[6]),'liners':self._s(row[7])}
                for i,k in enumerate(['accuracy','repeatability','turndown','min_cond','velocity','min_temp','max_temp','pressure_std','pressure_opt','pipe_sizes','excitation','protocols','epd','ip','transmitter','power','diagnostics','source','notes'],8):
                    d[k] = self._s(row[i]) if len(row)>i else ''
                models.append(d)
        return models
    def _load_drift(self):
        ws = self.wb['Drift Detection']; items = []
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            n = self._s(row[0])
            if n: items.append({'indicator':n,'desc':self._s(row[1]),'how':self._s(row[2]),'baseline':self._s(row[3]),'risk_rule':self._s(row[4]),'root':self._s(row[5]),'action':self._s(row[6]),'emerson':self._s(row[7]),'eh':self._s(row[8]),'krohne':self._s(row[9])})
        return items
    def _load_capex(self):
        ws = self.wb['CAPEX Components']; scores = {}
        for r in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            n, s = self._s(r[0]), r[2] if len(r)>2 else None
            if n and isinstance(s,(int,float)): scores[n] = int(s)
        return scores
    def _load_projects(self):
        ws = self.wb['JESA Project Data']; projects = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            p = self._s(row[0]); f = self._s(row[2]) if len(row)>2 else ''
            if p and f and not any(k in p for k in ['LEGEND','NOTE','NOTES']):
                projects.append({'project':p,'tag':self._s(row[1]),'fluid':f,'type':self._s(row[3]) if len(row)>3 else '','electrode':self._s(row[6]) if len(row)>6 else '','liner':self._s(row[7]) if len(row)>7 else '','grounding':self._s(row[8]) if len(row)>8 else '','tube':self._s(row[13]) if len(row)>13 else ''})
        return projects
    def get_asme_pressure(self, cls, temp):
        if cls not in self.asme: return None
        t = self.asme[cls]; temps = sorted(t.keys())
        for tp in temps:
            if temp <= tp: return t[tp]
        return t[temps[-1]]
    def get_liner_dn_limit(self, liner, vendor):
        for lk, vendors in self.liner_dn.items():
            if lk.upper().replace(' ','') in liner.upper().replace(' ','') or liner.upper().replace(' ','') in lk.upper().replace(' ',''):
                return vendors.get(vendor)
        return None

@dataclass
class ProcessInput:
    fluid_name:str; pipe_material:str; pipe_thickness:float=0.0; pipe_liner:str='Not applicable'
    tube_material:str='SS 316L'; dn:int=80; flow_normal:float=0.0; flow_max:float=0.0
    temp_min:float=0.0; temp_normal:float=0.0; temp_max:float=0.0; temp_design:float=0.0
    pressure_min:float=0.0; pressure_normal:float=0.0; pressure_max:float=0.0; pressure_design:float=0.0
    pressure_drop:float=0.0; conductivity:float=5.0; viscosity:float=1.0; density:float=1000.0
    special_conditions:str=''; user_notes:str=''

@dataclass
class ValidationResult:
    is_valid:bool=True; fluid_found:bool=False; fluid_data:Optional[Dict]=None
    fluid_category:str=''; velocity:float=0.0; recommended_dn:int=0
    errors:List[str]=field(default_factory=list); warnings:List[str]=field(default_factory=list)

@dataclass
class MaterialResult:
    electrode:str=''; electrode_alt:List[str]=field(default_factory=list); electrode_cost:str=''; electrode_avail:str='Standard'
    liner:str=''; liner_alt:List[str]=field(default_factory=list); liner_cost:str=''; liner_avail:str='Standard'; liner_dn_warning:str=''
    grounding:str=''; penetrant:str=''; o_ring:str=''; flange_coat:str=''; tube_warning:str=''
    warnings:List[str]=field(default_factory=list); remarks:str=''

@dataclass
class VendorRec:
    vendor:str=''; model:str=''; accuracy:str=''; pressure:str=''; protocols:str=''; diagnostics:str=''
    pipe_sizes:str=''; excitation:str=''; ip:str=''; notes:str=''; capex_score:int=3
    compatible:bool=True; reason:str=''; model_type:str=''; special_order:bool=False

@dataclass
class DriftRisk:
    indicator:str=''; level:str=''; description:str=''; steps:List[str]=field(default_factory=list)
    tools:Dict[str,str]=field(default_factory=dict); frequency:str=''

@dataclass
class ProjectMatch:
    project:str=''; fluid:str=''; match_type:str=''; confidence:int=0
    details:Dict[str,str]=field(default_factory=dict); explanation:str=''; cost_impact:str=''

@dataclass
class MaintenancePlan:
    continuous:List[str]=field(default_factory=list); monthly:List[str]=field(default_factory=list)
    quarterly:List[str]=field(default_factory=list); semi_annual:List[str]=field(default_factory=list)
    annual:List[str]=field(default_factory=list); multi_year:List[str]=field(default_factory=list)
    replacement:List[str]=field(default_factory=list)

@dataclass
class TCOResult:
    capex_score:int=0; breakdown:Dict=field(default_factory=dict); calib_months:int=24
    liner_life:str=''; electrode_life:str=''; drift_risks:List[DriftRisk]=field(default_factory=list)
    maintenance:MaintenancePlan=field(default_factory=MaintenancePlan)
    validation:List[ProjectMatch]=field(default_factory=list); notes_response:str=''; pressure_check:str=''

def mat_match(needle, haystack):
    n, h = needle.lower().strip(), haystack.lower()
    if not n or not h: return False
    if n in h: return True
    aliases = {'monel 400':['monel'],'monel':['monel 400'],'hastelloy c-4':['hastelloy c4','hast c-4'],'hastelloy c-276':['c-276','c276'],'hastelloy c':['hastelloy'],'alloy c22':['c22','alloy c-22'],'conductive rubber':['conductive'],'platinum':['platinum','pt-ir','pt/ir','80%pt'],'platinum / pt-ir':['platinum'],'soft rubber':['soft rubber','natural rubber'],'hard rubber':['hard rubber','ebonite'],'pfa':['pfa'],'ptfe':['ptfe'],'neoprene':['neoprene'],'linatex':['linatex'],'ceramic':['ceramic','al2o3'],'polyurethane':['polyurethane','pu'],'316l':['316l','316'],'tantalum':['tantalum'],'titanium':['titanium'],'tungsten carbide':['tungsten'],'316 ti':['316 ti','316ti']}
    for a in aliases.get(n,[]) + [n]:
        if a in h: return True
    return False

def get_cost(capex, mat):
    cm = {1:'Very Low',2:'Low',3:'Medium',4:'High',5:'Very High',6:'Extremely High'}
    for name, score in capex.items():
        if mat_match(mat, name): return cm.get(score,'?'), score
    return '?', 3

def get_category(ft):
    ft = ft.lower()
    if 'abrasive' in ft: return 'Abrasive'
    if any(k in ft for k in ['corrosive','acid']): return 'Corrosive'
    return 'Clean'

def is_conductive(pipe_mat, pipe_liner):
    if pipe_mat.upper() in ['FRP','PVC','HDPE','GRP']: return False
    if pipe_liner.lower() in ['hdpe','rubber lined']: return False
    return True

def calc_dn(flow_max, max_v=10.0):
    if flow_max <= 0: return 0
    d = math.sqrt(4*(flow_max/3600.0/max_v)/math.pi)*1000
    for dn in [15,25,40,50,65,80,100,125,150,200,250,300,350,400,450,500,600,700,800,900,1000,1200,1400,1600,1800,2000,2400,3000]:
        if dn >= d: return dn
    return 3000

class Layer1:
    def __init__(self, data): self.data = data
    def validate(self, inp):
        r = ValidationResult()
        fluid = self._find(inp.fluid_name)
        if fluid: r.fluid_found=True; r.fluid_data=fluid; r.fluid_category=get_category(fluid.get('type',''))
        else: r.warnings.append(f"Fluid '{inp.fluid_name}' not in database. Closest: {self._suggest(inp.fluid_name)}"); r.fluid_category='Unknown'
        if inp.conductivity < 5.0: r.is_valid=False; r.errors.append(f"Conductivity {inp.conductivity} µS/cm < 5 µS/cm. Magflow NOT suitable.")
        if inp.dn > 0 and inp.flow_max > 0:
            area = math.pi*(inp.dn/2000.0)**2; r.velocity = round((inp.flow_max/3600.0)/area, 2)
            is_slurry = r.fluid_category == 'Abrasive'
            if is_slurry:
                if r.velocity > 6.0: r.is_valid=False; r.errors.append(f"Velocity {r.velocity} m/s > 6 m/s max for slurries.")
                elif r.velocity > 3.0: r.warnings.append(f"Velocity {r.velocity} m/s > 3 m/s recommended for slurries.")
            elif r.velocity > 10.0: r.is_valid=False; r.errors.append(f"Velocity {r.velocity} m/s > 10 m/s max.")
        max_v = 3.0 if r.fluid_category=='Abrasive' else 10.0
        r.recommended_dn = calc_dn(inp.flow_max, max_v)
        if r.recommended_dn > 0 and inp.dn > 0 and inp.dn < r.recommended_dn:
            r.warnings.append(f"Recommended DN: {r.recommended_dn} mm. Selected DN {inp.dn} may cause high velocity.")
        if inp.dn < 2.5: r.is_valid=False; r.errors.append("DN below 2.5 mm minimum.")
        if inp.dn > 3000: r.is_valid=False; r.errors.append("DN exceeds 3000 mm maximum.")
        if fluid and 'demin' in fluid.get('name','').lower(): r.warnings.append("Demineralized water may drop below 5 µS/cm.")
        return r
    def _find(self, name):
        nl = name.lower().strip()
        for f in self.data.fluids:
            if f['name'].lower() == nl: return f
        for f in self.data.fluids:
            if nl in f['name'].lower() or f['name'].lower() in nl: return f
        return None
    def _suggest(self, name):
        nl = name.lower(); m = sorted([(len(set(nl.split())&set(f['name'].lower().split())),f['name']) for f in self.data.fluids], reverse=True)
        return ', '.join(x[1] for x in m[:3] if x[0]>0) or 'None'

class Layer2:
    def __init__(self, data): self.data = data
    def select(self, inp, fluid):
        r = MaterialResult()
        opts = [e.strip() for e in fluid.get('electrode','').replace(' ou ','/').replace(' or ','/').split('/') if e.strip()]
        r.electrode = opts[0] if opts else ''; r.electrode_alt = opts[1:] if len(opts)>1 else []
        if len(opts)>1: r.warnings.append(f"Multiple electrode options: {', '.join(opts)}. Depends on client requirements.")
        r.electrode_cost, _ = get_cost(self.data.capex_scores, r.electrode)
        for e in self.data.electrodes:
            if mat_match(r.electrode, e['name']): r.electrode_avail = e.get('availability','Standard'); break
        dt = inp.temp_design if inp.temp_design>0 else inp.temp_max if inp.temp_max>0 else inp.temp_normal
        liner_opts = [l.strip() for l in fluid.get('liner','').replace(' ou ','/').replace(' or ','/').split('/') if l.strip()]
        if len(liner_opts)>1: r.liner = self._pick_liner(liner_opts, dt, inp.dn); r.liner_alt = [l for l in liner_opts if l!=r.liner]
        else: r.liner = liner_opts[0] if liner_opts else ''
        r.liner_cost, _ = get_cost(self.data.capex_scores, r.liner)
        for li in self.data.liners:
            if mat_match(r.liner, li['name']): r.liner_avail = li.get('availability','Standard'); break
        for vn in ['E+H','Emerson','Krohne']:
            dl = self.data.get_liner_dn_limit(r.liner, vn)
            if dl and inp.dn > dl: r.liner_dn_warning += f"{r.liner} not available above DN {dl} at {vn}. "
        if r.liner_dn_warning: r.warnings.append(f"DN limitation: {r.liner_dn_warning}PTFE recommended for large DN.")
        cond = is_conductive(inp.pipe_material, inp.pipe_liner); gs = fluid.get('grounding','')
        r.grounding = gs if not cond and gs and 'strap' not in gs.lower() else ('Grounding strap (conductive pipe)' if cond else f"Required — non-conductive pipe ({inp.pipe_material})")
        r.penetrant = fluid.get('penetrant','N/A')
        if 'abrasive' in fluid.get('type','').lower() and r.penetrant in ['N/A','']: r.penetrant = 'Recommended (abrasive fluid)'; r.warnings.append("Abrasive fluid — penetrant ring recommended.")
        r.o_ring = fluid.get('o_ring',''); r.flange_coat = fluid.get('flange_coat',''); r.remarks = fluid.get('remarks','')
        if inp.tube_material in ['SS 304','SS 304L'] and get_category(fluid.get('type','')) in ['Corrosive','Abrasive']:
            r.tube_warning = f"{inp.tube_material} acceptable (liner protects tube). Cost saving with client approval."
        return r
    def _pick_liner(self, opts, temp, dn):
        if 'Soft rubber' in opts and temp <= 60: return 'Soft rubber'
        if any('PFA' in o for o in opts) and any('PTFE' in o for o in opts):
            if dn > 200: return 'PTFE'
            if temp > 150: return 'PFA HT' if temp <= 180 else 'PTFE'
            return 'PFA'
        return opts[0]

class Layer3:
    EXTRA = [{'vendor':'ABB','model':'ProcessMaster FEP630','accuracy':'±0.4%','notes':'Wide DN range, mining'},{'vendor':'Siemens','model':'SITRANS FM MAG 5100W','accuracy':'±0.4%','notes':'Water/wastewater'},{'vendor':'Yokogawa','model':'ADMAG AXF','accuracy':'±0.35%','notes':'Chemical process'},{'vendor':'Honeywell','model':'VersaFlow Mag 1000','accuracy':'±0.5%','notes':'Cost-effective'},{'vendor':'VEGA','model':'VEGAMAG 81','accuracy':'±0.3%','notes':'Compact'}]
    def __init__(self, data): self.data = data
    def select(self, inp, mat, fluid_data=None):
        compatible = []; cat = get_category((fluid_data or {}).get('type',''))
        for m in self.data.vendor_models:
            rec = VendorRec(vendor=m['vendor'],model=m['model'],accuracy=m['accuracy'],pressure=m['pressure_std'],protocols=m['protocols'],diagnostics=m['diagnostics'],pipe_sizes=m['pipe_sizes'],excitation=m['excitation'],ip=m['ip'],notes=m.get('notes',''),model_type=m.get('type',''))
            mt = m.get('type','').lower()
            if cat in ['Corrosive','Abrasive'] and any(w in mt for w in ['water','basic water','compact water','standard water','advanced water','premium water']): rec.compatible=False; rec.reason=f"Water model not for {cat}"
            if cat=='Clean' and any(k in mt for k in ['slurry','abrasive','aggressive']): rec.compatible=False; rec.reason="Slurry model not needed"
            if rec.compatible:
                ae = m.get('electrodes','')
                if not mat_match(mat.electrode, ae) and not any(mat_match(a, ae) for a in mat.electrode_alt):
                    if mat.electrode_avail.lower().startswith('special') and m['vendor']=='Krohne': rec.special_order = True
                    else: rec.compatible=False; rec.reason=f"Electrode {mat.electrode} not available"
            if rec.compatible:
                al = m.get('liners',''); lc = mat.liner.split('/')[0].strip()
                if not mat_match(lc, al) and not any(mat_match(a, al) for a in mat.liner_alt): rec.compatible=False; rec.reason=f"Liner {mat.liner} not available"
            if rec.compatible and not self._dn_ok(inp.dn, m.get('pipe_sizes','')): rec.compatible=False; rec.reason=f"DN {inp.dn} outside range"
            if rec.compatible:
                dt = inp.temp_design if inp.temp_design>0 else inp.temp_max
                try:
                    mt_v = float(m.get('max_temp','999').replace('°C','').replace('+','').strip())
                    if dt > mt_v: rec.compatible=False; rec.reason=f"Temp {dt}°C > max {mt_v}°C"
                except: pass
            if rec.compatible and inp.pressure_design > 0:
                cls = self._extract_class(m.get('pressure_std',''))
                if cls:
                    dt2 = inp.temp_design if inp.temp_design>0 else inp.temp_max if inp.temp_max>0 else 20
                    max_p = self.data.get_asme_pressure(cls, dt2)
                    if max_p and inp.pressure_design > max_p: rec.compatible=False; rec.reason=f"Pressure {inp.pressure_design} bar > {max_p} bar (Class {cls} at {dt2}°C)"
            rec.capex_score = self.data.capex_scores.get(m['model'], 3)
            if rec.compatible: compatible.append(rec)
        result = {'all':compatible,'emerson':None,'eh':None,'krohne':None,'extra':self.EXTRA,'warnings':[]}
        for vk, attr in [('Emerson','emerson'),('Endress+Hauser','eh'),('Krohne','krohne')]:
            vm = [c for c in compatible if c.vendor==vk]
            if vm:
                def score(r):
                    s = r.capex_score; mn = r.model.lower()
                    if cat=='Abrasive':
                        if any(k in mn for k in ['slurry','5100','5300','8707']): s-=10
                        elif any(k in mn for k in ['process','p 300','p 500','4100','4300','4400','8705']): s-=5
                    elif cat=='Corrosive':
                        if any(k in mn for k in ['p 300','p 500','4300','4400','8705']): s-=5
                        elif any(k in mn for k in ['process','p 100','p 10','4100']): s-=3
                    elif cat=='Clean':
                        if any(k in mn for k in ['w 400','w 500','w 300','2100','2300','8750']): s-=5
                    if r.special_order: s+=10
                    return s
                vm.sort(key=score); result[attr] = vm[0]
            else: result['warnings'].append(f"No compatible {vk} model found.")
        return result
    def _dn_ok(self, dn, dr):
        if not dr or dr=='N/A': return True
        try:
            c = dr.replace('mm','').replace('DN','').replace(' ',''); sep = '–' if '–' in c else '-'; p = c.split(sep); return float(p[0])<=dn<=float(p[1])
        except: return True
    def _extract_class(self, ps):
        if not ps: return None
        ps = ps.lower()
        for cls in [2500,1500,900,600,300,150]:
            if str(cls) in ps: return cls
        if 'pn 40' in ps: return 300
        if 'pn 16' in ps: return 150
        return None

class Layer4:
    def __init__(self, data): self.data = data
    def analyze(self, inp, mat, vendor, fluid):
        r = TCOResult(); ft = fluid.get('type','').lower(); cat = get_category(ft)
        dt = inp.temp_design if inp.temp_design>0 else inp.temp_max; fn = fluid.get('name','')
        _,es = get_cost(self.data.capex_scores, mat.electrode); _,ls = get_cost(self.data.capex_scores, mat.liner)
        vs = vendor.capex_score if vendor else 3; gs = 1 if 'strap' in mat.grounding.lower() else 3
        r.capex_score = es+ls+vs+gs+1; r.breakdown = {'Electrode':es,'Liner':ls,'Model':vs,'Grounding':gs,'Accessories':1}
        r.calib_months = 6 if cat=='Abrasive' else (12 if cat=='Corrosive' else 24)
        life = {'Abrasive':('3-5 years','5-10 years'),'Corrosive':('8-12 years','12-18 years'),'Clean':('15-20 years','20+ years')}
        r.liner_life, r.electrode_life = life.get(cat, ('10-15 years','15-20 years'))
        if vendor and inp.pressure_design > 0:
            cls = self._extract_class(vendor.pressure)
            if cls:
                dt2 = dt if dt > 0 else 20; max_p = self.data.get_asme_pressure(cls, dt2)
                if max_p: r.pressure_check = f"Class {cls} at {dt2}°C: max {max_p} bar. Design {inp.pressure_design} bar: {'OK ✓' if inp.pressure_design <= max_p else 'EXCEEDS ✗'}"
        r.drift_risks = self._drift(cat, inp, mat, vendor, fn, dt)
        r.maintenance = self._maint(cat, r.calib_months, vendor, mat)
        r.validation = self._validate(inp, mat, fluid)
        if inp.special_conditions or inp.user_notes: r.notes_response = self._notes(inp, mat, vendor, cat)
        return r
    def _extract_class(self, ps):
        if not ps: return None
        ps = ps.lower()
        for cls in [2500,1500,900,600,300,150]:
            if str(cls) in ps: return cls
        if 'pn 40' in ps: return 300
        if 'pn 16' in ps: return 150
        return None
    def _drift(self, cat, inp, mat, vendor, fn, dt):
        risks = []
        lvl = 'High' if cat=='Abrasive' else ('Medium' if cat=='Corrosive' else 'Low')
        desc_map = {'High':f"For {fn} with {mat.electrode} at {dt}°C, risk is HIGH. Abrasive particles cause wear and fouling.",'Medium':f"For {fn} with {mat.electrode} at {dt}°C, risk is MEDIUM. Long-term exposure causes gradual surface changes.",'Low':f"For {fn} with {mat.electrode} at {dt}°C, risk is LOW. Clean water causes minimal fouling."}
        freq_map = {'High':'Every 6 months','Medium':'Every 12 months','Low':'Every 24 months (extendable to 48 with SMV/Heartbeat)'}
        risks.append(DriftRisk('Electrode resistance',lvl, desc_map[lvl],['1. Check electrode resistance via HART — compare to baseline',f'2. If deviation <{"10" if lvl!="High" else "5"}% — no action',f'3. If deviation {"10-20" if lvl!="High" else "5-15"}% — schedule inspection',f'4. If deviation {"20-50" if lvl!="High" else "15-30"}% — plan replacement','5. If >50% or open circuit — URGENT replacement'],{'Emerson':'Smart Meter Verification','E+H':'Heartbeat Verification','Krohne':'OPTICHECK'},freq_map[lvl]))
        if cat in ['Abrasive','Charged']:
            risks.append(DriftRisk('Process noise','High',f"Abrasive particles in {fn} cause high-frequency noise.",['1. Monitor noise continuously','2. If >2x baseline — check solids/velocity','3. If >5x — increase damping, verify ≤3 m/s','4. Consider dual-frequency (Krohne IFC 300)','5. Verify penetrant rings intact'],{'Emerson':'DA1 High Process Noise Detection','E+H':'Heartbeat Monitoring — noise trending','Krohne':'Dual-frequency on IFC 300/400'},'Continuous'))
        if cat in ['Abrasive','Corrosive']:
            ll = 'High' if cat=='Abrasive' else 'Medium'
            risks.append(DriftRisk('Liner integrity',ll,f"{mat.liner} liner in {fn} at {dt}°C will degrade. Estimated life: {self._ll(mat.liner,cat)}.",['1. Monitor process noise trend','2. Visual inspection during shutdowns','3. Check liner at flange face','4. If damage confirmed — replace sensor','5. Preventive replacement based on lifespan'],{'Emerson':'Indirect via process noise','E+H':'Heartbeat — build-up index','Krohne':'Indirect via conductivity diagnostics'},f"Every {'6' if cat=='Abrasive' else '12'} months"))
        if not is_conductive(inp.pipe_material, inp.pipe_liner):
            risks.append(DriftRisk('Grounding fault','Medium',f"Non-conductive pipe ({inp.pipe_material}) — grounding ring integrity critical.",['1. Measure ground impedance — must be <1 Ohm','2. If 1-10 Ohm — clean contacts','3. If >10 Ohm — replace grounding rings','4. Check bolt tightness','5. Alternative: Krohne Virtual Reference (IFC 300)'],{'Emerson':'Grounding/Wiring Fault diagnostic','E+H':'Ground impedance <1 Ohm spec','Krohne':'Virtual Reference on IFC 300'},'Every 12 months'))
        risks.append(DriftRisk('Zero drift','Low',"Normal over time. Zero point shifts due to residue, installation, or temperature.",['1. Close block valves (full pipe, zero flow)','2. Wait for fluid to settle','3. Run auto-zero (~90 seconds)','4. Verify within ±0.5 mm/s of zero','5. If cannot restore — investigate contamination'],{'Emerson':'Auto-zero (~90s)','E+H':'Auto-zero + Heartbeat','Krohne':'Auto-zero all IFC'},f"Every {12 if cat in ['Corrosive','Abrasive'] else 24} months"))
        return risks
    def _ll(self, liner, cat):
        l = liner.lower()
        if 'soft rubber' in l: return '3-5 years'
        if 'ceramic' in l: return '10-15 years'
        if 'pfa' in l: return '5-8 years' if cat=='Abrasive' else '10-15 years'
        if 'ptfe' in l: return '3-5 years' if cat=='Abrasive' else '8-12 years'
        return '10-15 years'
    def _maint(self, cat, cm, vendor, mat):
        m = MaintenancePlan(); d = vendor.diagnostics.lower() if vendor else ''
        m.continuous = ['Transmitter self-test (automatic)','Empty pipe detection (continuous)']
        if 'heartbeat' in d: m.continuous.append('Heartbeat Monitoring — electrode/noise trending')
        elif 'smart meter' in d: m.continuous.append('Smart Meter Verification — background diagnostics')
        m.monthly = ['Check HART/diagnostic alarms in DCS','Verify transmitter display — no error codes']
        m.quarterly = ['Visual inspection — corrosion, cables','Cable gland tightness — IP integrity','Junction box moisture check']
        if cat=='Abrasive': m.semi_annual = ['Liner inspection via noise trend','Penetrant ring check','Electrode resistance check','Re-zero if needed']
        m.annual = [f'Calibration verification (every {cm} months)','Grounding verification (<1 Ohm)','Electrode resistance vs baseline','O-ring inspection','Electrical connections check']
        if 'heartbeat' in d or 'smart meter' in d: m.annual.append(f'Extended calibration: every {cm*2} months with diagnostic verification')
        m.multi_year = ['O-ring replacement (every 3-5 years)','Full bench calibration (every 3-5 years)','Liner evaluation — visual inspection','Grounding ring condition check']
        m.replacement = [f'Liner: every {self._ll(mat.liner,cat)}',f'Electrode: every {self._el(cat)}','O-ring: every 3-5 years','Transmitter: 15-20 years']
        return m
    def _el(self, cat): return '5-10 years' if cat=='Abrasive' else ('12-18 years' if cat=='Corrosive' else '20+ years')
    def _validate(self, inp, mat, fluid):
        matches = []; fn = inp.fluid_name.lower()
        for p in self.data.project_data:
            pf = p.get('fluid','').lower()
            if fn in pf or pf in fn:
                det = {}; conf = 100; expl = ''; cost = ''
                pe, pl, pt = p.get('electrode',''), p.get('liner',''), p.get('tube','')
                if mat_match(mat.electrode, pe): det['Electrode'] = f"{pe} ✓"
                elif pe:
                    det['Electrode'] = f"Project: {pe} | Model: {mat.electrode}"; conf -= 20
                    c1,s1 = get_cost(self.data.capex_scores,pe); c2,s2 = get_cost(self.data.capex_scores,mat.electrode)
                    expl = f"Project used {pe} ({c1}), model recommends {mat.electrode} ({c2})."; cost = f"{pe} ({c1}) vs {mat.electrode} ({c2})"
                if mat_match(mat.liner, pl): det['Liner'] = f"{pl} ✓"
                elif pl: det['Liner'] = f"Project: {pl} | Model: {mat.liner}"; conf -= 10
                if pt and pt!='N/A':
                    if mat_match(inp.tube_material, pt): det['Tube'] = f"{pt} ✓"
                    else: det['Tube'] = f"Project: {pt} | Selected: {inp.tube_material}"; conf -= 5
                mt = '✓ MATCH' if conf>=80 else ('⚠ PARTIAL' if conf>=50 else '✗ DIFFERS')
                matches.append(ProjectMatch(p['project'],p['fluid'],mt,max(conf,0),det,expl,cost))
        matches.sort(key=lambda x:-x.confidence); return matches[:8]
    def _notes(self, inp, mat, vendor, cat):
        parts = []; c = (inp.special_conditions+' '+inp.user_notes).lower()
        if 'atex' in c or 'hazardous' in c: parts.append("ATEX: All 3 vendors offer Ex d/Ex ia certified versions.")
        if 'sil' in c: parts.append("SIL: Krohne OPTIFLUX 4400 = dedicated SIL 2/3. E+H P 200 also SIL approved.")
        if 'bidirectional' in c or 'reverse' in c: parts.append("Bi-directional: All magflows support this by default.")
        if 'outdoor' in c or 'coastal' in c: parts.append("Outdoor/Coastal: Use IP68. Consider SS 316L housing. Apply ISO 12944 C5-M.")
        if 'budget' in c or 'cost' in c: parts.append(f"Budget: Consider SS 304L tube, compact transmitter, PTFE instead of PFA.")
        if not parts: parts.append("Notes acknowledged. No specific rule triggered.")
        return '\n\n'.join(parts)

class MagFlowAI:
    def __init__(self, path):
        self.data = DataLoader(path)
        self.l1,self.l2,self.l3,self.l4 = Layer1(self.data),Layer2(self.data),Layer3(self.data),Layer4(self.data)
    def recommend(self, inp):
        v = self.l1.validate(inp)
        if not v.is_valid: return {'validation':v,'materials':None,'vendors':None,'tco':None}
        m = self.l2.select(inp, v.fluid_data) if v.fluid_data else MaterialResult()
        vendors = self.l3.select(inp, m, v.fluid_data)
        best = vendors['emerson'] or vendors['eh'] or vendors['krohne']
        tco = self.l4.analyze(inp, m, best, v.fluid_data or {})
        return {'validation':v,'materials':m,'vendors':vendors,'tco':tco}
    def get_fluid_names(self): return sorted(set(f['name'] for f in self.data.fluids))
    def get_fluid_category(self, name):
        for f in self.data.fluids:
            if f['name'].lower()==name.lower(): return get_category(f.get('type',''))
        return 'Unknown'

if __name__=='__main__':
    ai = MagFlowAI('Data_Collection_v2.xlsx')
    print(f"Loaded: {len(ai.data.fluids)} fluids, {len(ai.data.electrodes)} electrodes, {len(ai.data.liners)} liners, {len(ai.data.vendor_models)} models, {len(ai.data.drift_indicators)} drift, {len(ai.data.project_data)} projects")
    t = ProcessInput('Strong phosphoric acid','FRP',temp_design=105,pressure_design=8,dn=80,flow_normal=23,flow_max=26,conductivity=10000)
    r = ai.recommend(t)
    v = r['vendors']
    print(f"E: {r['materials'].electrode}, L: {r['materials'].liner}")
    print(f"Emerson: {v['emerson'].model if v['emerson'] else 'None'}, E+H: {v['eh'].model if v['eh'] else 'None'}, Krohne: {v['krohne'].model if v['krohne'] else 'None'}")
