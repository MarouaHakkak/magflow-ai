"""MagFlow AI v9 — Landing page, colorful UI, offline maintenance Excel, state-safe"""
import streamlit as st
import anthropic
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import io
import json
import html
import re
from difflib import SequenceMatcher
import gspread
from google.oauth2.service_account import Credentials
from magflow_ai import MagFlowAI, ProcessInput
import base64
from github import Github

GITHUB_REPO = "MarouaHakkak/magflow-ai"
EXCEL_FILE_PATH = "Data_Collection_v2.xlsx"

# ============================================================
#  IMAGE URLS  —  REPLACE THE PLACEHOLDERS BELOW
#  (See the chat for step-by-step instructions on getting a
#   direct image URL. It must end in .png / .jpg and load on
#   its own in a browser tab.)
# ============================================================
IMG_JESA_LOGO = "https://i.imgur.com/K0ObAF4.png"
IMG_HERO      = "https://media.licdn.com/dms/image/v2/C561BAQEdzlQbQ6QRaQ/company-background_10000/company-background_10000/0/1583908059030?e=2147483647&v=beta&t=mn-3EDPCiDBT3qiDScVJVLzAlN_XvD2F3rDBL1YinXQ"
IMG_FLOWMETER = "https://upload.wikimedia.org/wikipedia/commons/2/24/%D0%A0%D0%B0%D1%81%D1%85%D0%BE%D0%B4%D0%BE%D0%BC%D0%B5%D1%80_%D1%8D%D0%BB%D0%B5%D0%BA%D1%82%D1%80%D0%BE%D0%BC%D0%B0%D0%B3%D0%BD%D0%B8%D1%82%D0%BD%D1%8B%D0%B9.jpg"  # magnetic flowmeter photo (Wikimedia Commons)

# Vendor website links (used in Layer 3 "Other Vendors" + main 3 cards)
VENDOR_LINKS = {
    "Emerson":          "https://www.emerson.com/en-us/automation/measurement-instrumentation/flow-measurement/magnetic-flow-meters",
    "Endress+Hauser":   "https://www.endress.com/en/field-instruments-overview/flow-measurement-product-overview/electromagnetic-flowmeter",
    "Endress Hauser":   "https://www.endress.com/en/field-instruments-overview/flow-measurement-product-overview/electromagnetic-flowmeter",
    "Krohne":           "https://krohne.com/en/products/flow-measurement/flowmeters/electromagnetic-flowmeters",
    "ABB":              "https://new.abb.com/products/measurement-products/flow/electromagnetic-flowmeters",
    "Siemens":          "https://www.siemens.com/global/en/products/automation/process-instrumentation/flow-measurement/electromagnetic.html",
    "Yokogawa":         "https://www.yokogawa.com/solutions/products-and-services/measurement/flow-meters/magnetic-flow-meters/",
    "Honeywell":        "https://process.honeywell.com/us/en/products/field-instruments/flow-measurement",
    "VEGA":             "https://www.vega.com/en-us/products/product-catalog/flow",
}

def vendor_url(name):
    if not name:
        return None
    key = name.strip()
    if key in VENDOR_LINKS:
        return VENDOR_LINKS[key]
    for k, v in VENDOR_LINKS.items():
        if k.lower() in key.lower() or key.lower() in k.lower():
            return v
    return None

st.set_page_config(page_title="MagFlow AI — JESA", page_icon="🔧", layout="wide")

# ============================================================
#  GLOBAL STYLING  —  colorful theme + spaced tabs
# ============================================================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
:root {
    --jesa-blue: #0B55D9;
    --jesa-navy: #102A63;
    --soft-blue: #EEF6FF;
    --ink: #1F2937;
}
html, body, [class*="css"] {
    font-family: 'Inter', system-ui, -apple-system, BlinkMacSystemFont, sans-serif;
}
.block-container {
    padding-top: 1.1rem;
    padding-bottom: 2rem;
    max-width: 1480px;
}
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #FFFFFF 0%, #F7FBFF 100%);
    border-right: 1px solid #E6EEF8;
}
section[data-testid="stSidebar"] > div {
    padding-top: 1.4rem;
}
section[data-testid="stSidebar"] img {
    margin: 0 auto 1.3rem auto;
    display: block;
}
section[data-testid="stSidebar"] hr {
    margin: 1.1rem 0;
}
section[data-testid="stSidebar"] h3 {
    color: var(--ink);
    font-weight: 800;
    letter-spacing: -0.02em;
}
.mf-assistant-greeting {
    display: flex;
    gap: 12px;
    align-items: flex-start;
    color: #2E3442;
    font-size: 18px;
    line-height: 1.45;
    font-weight: 500;
    margin: 18px 0 14px 0;
}
.mf-bot-badge {
    width: 38px;
    height: 38px;
    border-radius: 12px;
    background: #FFA31A;
    color: white;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    flex: 0 0 auto;
}
.mf-assistant-reply {
    max-height: 150px;
    overflow-y: auto;
    background: #FFFFFF;
    border: 1px solid #EEF2F7;
    border-radius: 14px;
    padding: 12px 14px;
    color: #344054;
    font-size: 13px;
    line-height: 1.45;
    margin: 10px 0;
}
.mf-assistant-label {
    color: #7A8190;
    font-size: 15px;
    margin: 18px 0 8px 0;
}
/* Tab bar: bigger, clearer labels spread across the full page width (note 13) */
.stTabs [data-baseweb="tab-list"] {
    gap: 10px;
    display: flex;
    width: 100%;
    background: rgba(255,255,255,0.75);
    border: 1px solid #E4ECF7;
    border-radius: 18px;
    padding: 8px;
    box-shadow: 0 10px 28px rgba(16,42,99,0.06);
}
.stTabs [data-baseweb="tab"] {
    flex: 1 1 0;
    justify-content: center;
    text-align: center;
    border-radius: 14px;
    color: #5C6B84;
    min-height: 52px;
}
.stTabs [data-baseweb="tab"] p {
    font-size: 16px !important;
    font-weight: 700 !important;
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, #EAF3FF 0%, #FFFFFF 100%);
    box-shadow: inset 0 -3px 0 var(--jesa-blue);
}
.stTabs [aria-selected="true"] p {
    color: var(--jesa-blue) !important;
}
/* Soft page background */
.stApp { background: radial-gradient(circle at 76% 8%, #E3F1FF 0%, transparent 34%), linear-gradient(180deg, #F8FBFF 0%, #EEF4FA 100%); color: var(--ink); }
/* Buttons */
.stButton button, .stDownloadButton button {
    border-radius: 16px;
    font-weight: 700;
    border: 1px solid #DDE8F6;
    min-height: 48px;
}
.stButton button[kind="primary"], .stDownloadButton button[kind="primary"] {
    background: linear-gradient(135deg, #0B55D9 0%, #149AD6 100%) !important;
    border: none !important;
    color: white !important;
    box-shadow: 0 12px 28px rgba(11,85,217,0.22);
}
button[data-testid="stBaseButton-primary"] {
    background: linear-gradient(135deg, #0B55D9 0%, #149AD6 100%) !important;
    border: none !important;
    color: white !important;
    box-shadow: 0 12px 28px rgba(11,85,217,0.22);
}
section[data-testid="stSidebar"] .stButton button {
    background: #EEF5FF;
    color: var(--jesa-blue);
    border: 0;
    justify-content: flex-start;
    padding-left: 24px;
    box-shadow: none;
}
div[data-baseweb="select"] > div, input, textarea {
    border-radius: 14px !important;
}
/* Landing hero card */
.mf-hero {
    border-radius: 26px;
    overflow: hidden;
    box-shadow: 0 18px 46px rgba(21,42,74,0.17);
    margin: 10px auto 18px auto;
    border: 1px solid rgba(226,236,248,0.9);
}
.mf-hero img {
    width: 100%;
    display: block;
    max-height: 560px;
    object-fit: cover;
    object-position: center;
}
.mf-card {
    background: #FFFFFF;
    border-radius: 14px;
    padding: 22px;
    box-shadow: 0 3px 14px rgba(21,42,74,0.08);
    height: 100%;
}
.mf-landing-head {
    position: relative;
    min-height: 310px;
    padding: 88px 24px 42px 24px;
    display: flex;
    align-items: center;
    justify-content: center;
    text-align: center;
    background: linear-gradient(135deg, rgba(255,255,255,0.86) 0%, rgba(239,248,255,0.92) 58%, rgba(225,242,255,0.88) 100%);
    border-bottom: 1px solid rgba(199,222,248,0.82);
    margin: -8px -10px 22px -10px;
}
.mf-landing-logo {
    position: absolute;
    left: 32px;
    top: 92px;
    width: 185px;
}
.mf-landing-title {
    font-family: 'Inter', Arial, system-ui, sans-serif;
    font-size: clamp(72px, 6.8vw, 112px);
    line-height: 0.98;
    font-weight: 780;
    letter-spacing: -0.05em;
    color: #06246B;
    margin: 0;
    text-align: center;
    text-shadow: 0 2px 0 rgba(255,255,255,0.9), 0 10px 22px rgba(8,37,107,0.08);
}
.mf-landing-title .ai {
    color: #14A8D8;
    letter-spacing: -0.025em;
}
.mf-landing-sub {
    font-size: clamp(20px, 1.65vw, 28px);
    color: #1565C0;
    margin: 30px auto 0 auto;
    font-weight: 600;
    line-height: 1.25;
    text-align: center;
    max-width: 1240px;
}
.mf-landing-meta {
    font-size: clamp(15px, 1.18vw, 20px);
    color: #888;
    margin: 24px auto 0 auto;
    text-align: center;
}
.mf-home-title {
    color: #1B2A4A;
    font-size: 30px;
    font-weight: 850;
    margin: 34px 0 12px 0;
    letter-spacing: -0.02em;
}
.mf-home-subtitle {
    color: #1565C0;
    font-size: 22px;
    font-weight: 800;
    margin: 26px 0 14px 0;
}
.mf-home-copy {
    font-size: 16px;
    color: #444;
    line-height: 1.65;
    max-width: 1120px;
}
.mf-step {
    background: #FFFFFF;
    border-left: 5px solid #1565C0;
    border-radius: 10px;
    padding: 16px 18px;
    margin-bottom: 12px;
    box-shadow: 0 2px 10px rgba(21,42,74,0.06);
}
.mf-step b { color: #1565C0; }
.mf-app-hero {
    position: relative;
    overflow: hidden;
    min-height: 430px;
    border-radius: 0;
    padding: 0;
    margin: -6px -10px 30px -10px;
    background:
        radial-gradient(circle at 13% 58%, rgba(214,235,255,0.72) 0 10%, transparent 23%),
        radial-gradient(circle at 92% 7%, rgba(202,231,247,0.60) 0 9%, transparent 24%),
        linear-gradient(155deg, rgba(255,255,255,0.96) 0%, rgba(247,251,255,0.96) 46%, rgba(225,241,255,0.95) 100%);
    border-bottom: 1px solid rgba(190,216,246,0.85);
    box-shadow: 0 18px 46px rgba(16,42,99,0.08);
}
.mf-app-hero::before {
    content: "";
    position: absolute;
    left: -2%;
    top: 28%;
    width: 28%;
    height: 34%;
    background:
        radial-gradient(circle at 50% 50%, rgba(255,255,255,0.88) 0 29%, transparent 30%),
        linear-gradient(90deg, rgba(17,96,224,0.12), rgba(29,180,215,0.12));
    border-radius: 50%;
    filter: blur(0.2px);
    opacity: 0.95;
}
.mf-app-hero::after {
    content: "";
    position: absolute;
    right: -7%;
    bottom: -20%;
    width: 52%;
    height: 62%;
    background:
        linear-gradient(145deg, transparent 0 31%, rgba(207,229,253,0.58) 32% 54%, rgba(186,218,251,0.52) 55% 100%);
    border-radius: 60% 0 0 0;
    transform: skewY(-8deg);
}
.mf-hero-inner {
    position: relative;
    z-index: 1;
    display: flex;
    align-items: center;
    justify-content: center;
    text-align: center;
    width: 100%;
    min-height: 430px;
    margin: 0 auto;
    padding: 58px 60px 82px 60px;
}
.mf-wave-logo {
    display: none;
}
.mf-wave-logo::before {
    content: none;
}
.mf-wave-logo::after {
    content: none;
}
.mf-app-title {
    margin: 0;
    font-family: 'Inter', 'Arial', system-ui, sans-serif;
    font-size: clamp(70px, 7.3vw, 112px);
    line-height: 1;
    font-weight: 800;
    letter-spacing: -0.045em;
    color: #06246B;
    text-align: center;
    text-shadow: 0 2px 0 rgba(255,255,255,0.85), 0 10px 24px rgba(8,37,107,0.10);
}
.mf-app-title .ai {
    color: #14A8D8;
    letter-spacing: -0.025em;
}
.mf-app-sub {
    color: #40506A;
    font-size: clamp(18px, 1.45vw, 25px);
    font-weight: 600;
    line-height: 1.25;
    margin: 52px auto 26px auto;
    text-align: center;
    max-width: 1280px;
    white-space: normal;
}
.mf-app-meta {
    color: #6E7C92;
    font-size: clamp(16px, 1.28vw, 22px);
    font-weight: 500;
    margin: 0 auto;
    text-align: center;
}
.mf-home-row {
    margin: -18px 0 18px 0;
}
.mf-home-row div[data-testid="stVerticalBlock"],
.mf-home-row div[data-testid="stElementContainer"] {
    width: 100% !important;
}
.mf-app-home-shell {
    position: relative;
    z-index: 5;
}
.mf-home-row .stButton button {
    background: rgba(255,255,255,0.94) !important;
    color: #344054 !important;
    border: 1px solid #E2EAF5 !important;
    border-radius: 22px !important;
    box-shadow: 0 14px 34px rgba(16,42,99,0.12) !important;
    min-height: 54px;
    font-size: 16px;
    justify-content: center;
}
.mf-section-shell {
    background: rgba(255,255,255,0.72);
    border: 1px solid #E2ECF8;
    border-radius: 22px;
    padding: 18px;
    box-shadow: 0 14px 36px rgba(16,42,99,0.06);
    margin-top: 18px;
}
</style>
""", unsafe_allow_html=True)

@st.cache_resource
def load_model():
    return MagFlowAI('Data_Collection_v2.xlsx')

ai = load_model()

SHEET_ID = "1jVpJkHPtG808WtlKxKpcIxvNLvLyx8syIi0GlppNTgU"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets",
          "https://www.googleapis.com/auth/drive"]

def _norm_tag(tag):
    """Normalize tag numbers for matching while keeping stored/displayed values readable."""
    return "".join(str(tag or "").upper().split())

@st.cache_resource
def get_gsheet():
    try:
        creds_dict = dict(st.secrets["gcp_service_account"])
        creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
        client = gspread.authorize(creds)
        sheet = client.open_by_key(SHEET_ID).sheet1
        return sheet
    except Exception as e:
        return None

def load_history(tag_filter=None):
    sheet = get_gsheet()
    if sheet is None:
        return []
    try:
        records = sheet.get_all_records()
        if tag_filter:
            target_tag = _norm_tag(tag_filter)
            records = [r for r in records if _norm_tag(r.get('Tag Number','')) == target_tag]
        return records
    except:
        return []

def save_history(date, tag, intervention_type, task, result, technician, notes):
    sheet = get_gsheet()
    if sheet is None:
        return False
    try:
        sheet.append_row([date, str(tag).strip(), intervention_type, task, result, technician, notes])
        return True
    except:
        return False

# ============================================================
#  CHECKLIST PERSISTENCE  (pre-fill the checklist for a tag)
#  Stores filled checklist cells in a Google Sheet tab "Checklist Records"
#  Key per cell: tag :: period :: task :: column_label
# ============================================================
CHECKLIST_WS = "Checklist Records"

def _checklist_key(tag, period, task, col_label):
    return f"{_norm_tag(tag)}::{str(period).strip()}::{str(task).strip()}::{str(col_label).strip()}"

def _checklist_ws():
    """Return the gspread worksheet for checklist records (create if missing)."""
    try:
        creds_dict = dict(st.secrets["gcp_service_account"])
        creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
        client_gs = gspread.authorize(creds)
        ss = client_gs.open_by_key(SHEET_ID)
        try:
            ws = ss.worksheet(CHECKLIST_WS)
        except gspread.WorksheetNotFound:
            ws = ss.add_worksheet(title=CHECKLIST_WS, rows=2000, cols=6)
            ws.append_row(["Tag", "Period", "Task", "Column", "Value", "Updated"])
        return ws
    except Exception:
        return None

def load_checklist_records(tag):
    """Return {key: value} of saved checklist cells for this tag."""
    ws = _checklist_ws()
    if ws is None:
        return {}
    try:
        out = {}
        target_tag = _norm_tag(tag)
        for r in ws.get_all_records():
            if _norm_tag(r.get("Tag", "")) != target_tag:
                continue
            key = _checklist_key(r.get("Tag", ""), r.get("Period", ""), r.get("Task", ""), r.get("Column", ""))
            out[key] = str(r.get("Value", ""))
        return out
    except Exception:
        return {}

def save_checklist_records(tag, cells):
    """cells = {key: value}. Last-write-wins: append rows; latest read wins on load
    because get_all_records keeps order and we overwrite in the dict."""
    ws = _checklist_ws()
    if ws is None:
        return 0
    try:
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        rows = []
        for key, val in cells.items():
            parts = key.split("::")
            if len(parts) != 4:
                continue
            t, period, task, col = parts
            rows.append([t, period, task, col, val, now])
        if rows:
            ws.append_rows(rows, value_input_option="RAW")
        return len(rows)
    except Exception:
        return 0

def read_checklist_from_excel(uploaded_file, tag):
    """Read the 'Maintenance Checklist' sheet of an uploaded Excel and return
    {key: value} for every filled data cell. Re-derives period from headers."""
    try:
        wb = openpyxl.load_workbook(uploaded_file)
        if "Maintenance Checklist" not in wb.sheetnames:
            return {}
        ws = wb["Maintenance Checklist"]
        out = {}
        current_month = None
        current_period = None
        mode = None  # continuous / dated
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            a_str = str(a).strip() if a is not None else ""
            if a_str.startswith("📅"):
                toks = a_str.replace("📅", "").strip().split()
                current_month = toks[0] if toks else None
                continue
            if a_str.lstrip().startswith("Week"):
                wk = a_str.strip()
                current_period = f"{current_month}/{wk}"
                mode = "continuous"
                continue
            if "Monthly tasks" in a_str:
                current_period = f"{current_month}/Monthly"
                mode = "dated"
                continue
            if "Quarterly" in a_str and "—" in a_str:
                qlabel = a_str.split("—")[-1].strip()
                current_period = f"{current_month}/Quarterly {qlabel}"
                mode = "dated"
                continue
            if "Semi-annual" in a_str and "—" in a_str:
                slabel = a_str.split("—")[-1].strip()
                current_period = f"{current_month}/Semi-annual {slabel}"
                mode = "dated"
                continue
            if "ANNUAL MAINTENANCE" in a_str:
                current_period = "ANNUAL"
                mode = "dated"
                continue
            if "MULTI-YEAR" in a_str:
                current_period = "MULTI-YEAR"
                mode = "dated"
                continue
            if "COMPONENT REPLACEMENT" in a_str:
                current_period = "REPLACEMENT"
                mode = "dated"
                continue
            task = ws.cell(row=r, column=2).value
            is_task_row = bool(task and current_period and (a_str == "☐" or a_str.isdigit() or a_str == "•"))
            if is_task_row:
                task = str(task).strip()
                if mode == "continuous":
                    cols = [(3,"Mon"),(4,"Tue"),(5,"Wed"),(6,"Thu"),(7,"Fri"),(8,"Sat"),(9,"Sun"),
                            (10,"Done?"),(11,"Result"),(12,"Technician"),(15,"Validation")]
                elif current_period in ("ANNUAL", "MULTI-YEAR", "REPLACEMENT"):
                    cols = [(10,"Done?"),(11,"Work done"),(15,"Validation")]
                else:
                    cols = [(10,"Done?"),(11,"Result"),(12,"Technician"),(13,"Date"),(14,"Work done"),(15,"Validation")]
                for col, lab in cols:
                    val = ws.cell(row=r, column=col).value
                    if val is not None and str(val).strip() != "":
                        out[_checklist_key(tag, current_period, task, lab)] = str(val)
        return out
    except Exception:
        return {}

import calendar as _calendar

_MONTHS_IDX = {m:i for i,m in enumerate(
    ["January","February","March","April","May","June","July","August",
     "September","October","November","December"], 1)}

def _is_nonconform(val):
    if not val:
        return False
    v = str(val).strip().lower()
    return v in ("nc","non-conform","nonconform","non conform","not ok","ko","fail","failed","non conforme","non-conforme")

def _is_positive_result(val):
    if not val:
        return False
    v = str(val).strip().lower()
    return v in ("ok", "conform", "yes", "done", "dry", "within tolerance", "complete", "<1 ohm")

def _is_done(val):
    if not val:
        return False
    return str(val).strip().lower() in ("yes", "oui", "done", "fait", "x", "true", "1")

def _week_date(month_name, week_num, year):
    mi = _MONTHS_IDX.get(month_name)
    if not mi:
        return ""
    return f"{year}-{mi:02d} W{int(week_num)}"

def extract_interventions_from_checklist(uploaded_file, tag, year=None):
    """Read the filled checklist and return a list of intervention dicts.
    Any task marked Done? = Yes becomes an intervention. Older files with the
    previous checkbox/status layout are still supported."""
    if year is None:
        year = datetime.now().year
    try:
        wb = openpyxl.load_workbook(uploaded_file)
        if "Maintenance Checklist" not in wb.sheetnames:
            return []
        ws = wb["Maintenance Checklist"]
        interventions = []
        current_month = None
        current_period = None
        mode = None
        week_num = 1
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            a_str = str(a).strip() if a is not None else ""
            if a_str.startswith("📅"):  # calendar emoji month header
                toks = a_str.replace("📅", "").strip().split()
                current_month = toks[0] if toks else None
                continue
            if a_str.lstrip().startswith("Week"):
                mode = "continuous"
                try: week_num = int(a_str.lstrip().split()[1])
                except: week_num = 1
                current_period = f"{current_month}/Week {week_num}"
                continue
            if "Monthly tasks" in a_str:
                current_period = f"{current_month}/Monthly"
                mode = "dated"; continue
            if "Quarterly" in a_str and "—" in a_str:
                qlabel = a_str.split("—")[-1].strip()
                current_period = f"{current_month}/Quarterly {qlabel}"
                mode = "dated"; continue
            if "Semi-annual" in a_str and "—" in a_str:
                slabel = a_str.split("—")[-1].strip()
                current_period = f"{current_month}/Semi-annual {slabel}"
                mode = "dated"; continue
            if "ANNUAL MAINTENANCE" in a_str:
                current_period = "ANNUAL"
                mode = "simple"; continue
            if "MULTI-YEAR" in a_str:
                current_period = "MULTI-YEAR"
                mode = "simple"; continue
            if "COMPONENT REPLACEMENT" in a_str:
                current_period = "REPLACEMENT"
                mode = "simple"; continue
            task = ws.cell(row=r, column=2).value
            is_task_row = bool(task and (a_str == "☐" or a_str.isdigit() or a_str == "•"))
            if is_task_row:
                task = str(task).strip()
                new_format = a_str != "☐"
                if mode == "continuous":
                    status = ws.cell(row=r, column=11 if new_format else 10).value
                    tech = ws.cell(row=r, column=12 if new_format else 11).value
                    done = ws.cell(row=r, column=10).value if new_format else status
                    work = ws.cell(row=r, column=14).value if new_format else ""
                    day_marks = []
                    if new_format:
                        for col, day in [(3,"Mon"),(4,"Tue"),(5,"Wed"),(6,"Thu"),(7,"Fri"),(8,"Sat"),(9,"Sun")]:
                            val = ws.cell(row=r, column=col).value
                            if val is not None and str(val).strip():
                                day_marks.append(f"{day}: {val}")
                    has_continuous_data = any([
                        _is_done(done),
                        status is not None and str(status).strip(),
                        tech is not None and str(tech).strip(),
                        work is not None and str(work).strip(),
                        bool(day_marks),
                    ])
                    if has_continuous_data or (not new_format and _is_nonconform(status)):
                        period_note = f"Period: {current_month} Week {week_num}" if current_month else f"Week {week_num}"
                        notes = " | ".join([period_note, str(work or "").strip()] + day_marks).strip(" |")
                        interventions.append({
                            "date": _week_date(current_month, week_num, year),
                            "tag": tag, "type": "Checklist task", "task": task,
                            "result": str(status or "Conform").strip(), "tech": str(tech or "").strip(),
                            "notes": notes})
                elif mode == "simple":
                    done = ws.cell(row=r, column=10).value if new_format else None
                    notes = ws.cell(row=r, column=11).value if new_format else ""
                    if _is_done(done) or (notes and str(notes).strip()):
                        interventions.append({
                            "date": current_period or "",
                            "tag": tag, "type": "Checklist task", "task": task,
                            "result": "Done" if _is_done(done) else "Notes only", "tech": "",
                            "notes": str(notes or "").strip()})
                else:
                    done = ws.cell(row=r, column=10).value if new_format else None
                    result = ws.cell(row=r, column=11).value if new_format else "Conform"
                    tech = ws.cell(row=r, column=12 if new_format else 11).value
                    date = ws.cell(row=r, column=13 if new_format else 10).value
                    work = ws.cell(row=r, column=14 if new_format else 12).value
                    if _is_done(done) or (result and str(result).strip()) or (date and str(date).strip()) or (tech and str(tech).strip()) or (work and str(work).strip()):
                        interventions.append({
                            "date": str(date).strip() if date else "",
                            "tag": tag, "type": "Checklist task", "task": task,
                            "result": str(result or "Conform").strip(), "tech": str(tech or "").strip(),
                            "notes": str(work or "").strip()})
        return interventions
    except Exception:
        return []

def save_interventions_dedup(interventions):
    """Save interventions to history, skipping duplicates (same tag+date+task)."""
    existing = load_history()
    seen = set()
    for r in existing:
        seen.add((_norm_tag(r.get("Tag Number","")), str(r.get("Date","")).strip(), str(r.get("Task","")).strip()))
    added = 0
    for it in interventions:
        sig = (_norm_tag(it["tag"]), str(it["date"]).strip(), str(it["task"]).strip())
        if sig in seen:
            continue
        ok = save_history(it["date"], it["tag"], it["type"], it["task"], it["result"], it["tech"], it["notes"])
        if ok:
            seen.add(sig); added += 1
    return added

def import_from_excel(uploaded_file):
    try:
        wb = openpyxl.load_workbook(uploaded_file)
        if "Historical Data" not in wb.sheetnames:
            return 0, "Sheet 'Historical Data' not found in the uploaded file."
        ws = wb["Historical Data"]
        saved = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            date, tag, itype, task, result, tech, notes = (row[0], row[1], row[2], row[3], row[4], row[5], row[6] if len(row) > 6 else "")
            if not date or not tag or not task:
                continue
            success = save_history(str(date), str(tag), str(itype or ""), str(task), str(result or ""), str(tech or ""), str(notes or ""))
            if success:
                saved += 1
        return saved, None
    except Exception as e:
        return 0, str(e)

def build_context():
    fluids = ", ".join(f['name'] for f in ai.data.fluids[:25]) + f"... ({len(ai.data.fluids)} total)"
    return f"""You are MagFlow AI Assistant for JESA/OCP electromagnetic flowmeters.
Database: {len(ai.data.fluids)} fluids, {len(ai.data.electrodes)} electrodes, {len(ai.data.liners)} liners, {len(ai.data.vendor_models)} models, {len(ai.data.drift_indicators)} drift indicators, {len(ai.data.project_data)} project entries.
Fluids: {fluids}
Rules: PFA default for acids, PTFE for large DN, Soft rubber for rock slurry, Min 5 uS/cm, Max 10 m/s / 3-6 m/s slurries, Grounding ring for FRP/PVC.
Diagnostics: Emerson SMV, E+H Heartbeat, Krohne OPTICHECK.
Answer in same language as question. Source: [JESA Database], [General Knowledge], or [Web Search].
For new tech: USE web_search, provide URLs."""

def search_local(query):
    q = query.lower().replace('-',' '); results = []; words = [w for w in q.split() if len(w)>2]; combined = q.replace(' ','')
    for f in ai.data.fluids:
        fn = f['name'].lower(); fnc = fn.replace(' ','')
        if any(w in fn for w in words) or combined in fnc or fnc in combined:
            results.append(f"Fluid: {f['name']} | Type: {f['type']} | Electrode: {f['electrode']} | Liner: {f['liner']} | Grounding: {f['grounding']}")
    for e in ai.data.electrodes:
        if any(w in e['name'].lower() for w in words): results.append(f"Electrode: {e['name']} | Cost: {e['cost']}")
    for m in ai.data.vendor_models:
        if any(w in (m['model']+' '+m['vendor']).lower() for w in words): results.append(f"Model: {m['vendor']} {m['model']} | Accuracy: {m['accuracy']}")
    for d in ai.data.drift_indicators:
        if any(w in d['indicator'].lower() for w in words): results.append(f"Drift: {d['indicator']} | {d['desc']}")
    for p in ai.data.project_data:
        if any(w in (p.get('fluid','')+' '+p.get('project','')).lower() for w in words): results.append(f"Project: {p['project']} | {p['fluid']} | {p.get('electrode','')}")
    return results[:15]

def ask_claude(question, local_results, history=None):
    if history is None: history = []
    try:
        client = anthropic.Anthropic(api_key=st.secrets["ANTHROPIC_API_KEY"])
        ctx = ""
        if local_results: ctx = "\n\nJESA data:\n" + "\n".join(f"- {r}" for r in local_results)
        conv = ""
        if history:
            conv = "\n\nPrevious conversation:\n"
            for msg in history[-6:]:
                conv += f"{'User' if msg['role']=='user' else 'Assistant'}: {msg['content'][:200]}\n"
        response = client.messages.create(model="claude-haiku-4-5-20251001", max_tokens=1500, system=build_context(),
            tools=[{"type": "web_search_20250305", "name": "web_search"}],
            messages=[{"role": "user", "content": question + ctx + conv}])
        parts = [block.text for block in response.content if hasattr(block, 'text')]
        return "\n".join(parts) if parts else "No response."
    except Exception as e:
        return f"Error: {str(e)}"

def extract_datasheet_with_ai(pdf_bytes):
    """Extract instruments page by page to avoid JSON truncation."""
    try:
        from pypdf import PdfReader
        import io as _io
        from pypdf import PdfWriter

        client = anthropic.Anthropic(api_key=st.secrets["ANTHROPIC_API_KEY"])

        prompt = """This is ONE page from a JESA magnetic flowmeter datasheet.
If this page contains an instrument specification (has a Tag Number and process data), extract it as JSON.
If not (cover page, index, notes), return exactly: []
Return ONLY a JSON array with one object per instrument found:
[{"project":"...","tag":"...","service":"...","fluid":"...","dn":80,"flow_normal":23.0,"flow_max":26.0,"temp_design":105,"pressure_design":10.0,"conductivity":">5 uS/cm","electrode":"Tantalum","liner":"PFA","tube":"316L SS","grounding":"Grounding straps","accuracy":"0.2%","vendor":"VTA","model":"VTA"}]
Use the exact value from the PDF when it is present, but correct obvious OCR spelling only when the meaning is certain (example: PHOPHORIC ACID -> phosphoric acid).
Do not invent maintenance plans, CAPEX scores, or database ratings. Use null when the PDF does not provide a value.
Return ONLY JSON, no markdown."""

        reader = PdfReader(_io.BytesIO(pdf_bytes))
        all_instruments = []

        for page_num, page in enumerate(reader.pages):
            try:
                writer = PdfWriter()
                writer.add_page(page)
                page_buf = _io.BytesIO()
                writer.write(page_buf)
                page_buf.seek(0)
                page_b64 = base64.standard_b64encode(page_buf.read()).decode("utf-8")

                response = client.messages.create(
                    model="claude-haiku-4-5-20251001",
                    max_tokens=1000,
                    messages=[{"role": "user", "content": [
                        {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": page_b64}},
                        {"type": "text", "text": prompt}
                    ]}])

                text = response.content[0].text.strip().replace("```json","").replace("```","").strip()
                start = text.find('[')
                end = text.rfind(']')
                if start != -1 and end != -1:
                    page_instruments = json.loads(text[start:end+1])
                    all_instruments.extend(page_instruments)
            except:
                continue

        return all_instruments, None if all_instruments else ([], "No instruments found in any page of this PDF.")

    except Exception as e:
        return [], str(e)

def get_excel_from_github():
    try:
        token = st.secrets.get("GITHUB_TOKEN", None)
        if token is None:
            try: token = st.secrets["github"]["GITHUB_TOKEN"]
            except: pass
        if token is None:
            raise Exception("GITHUB_TOKEN not found in secrets")
        gh = Github(token)
        repo = gh.get_repo(GITHUB_REPO)
        contents = repo.get_contents(EXCEL_FILE_PATH)
        excel_bytes = base64.b64decode(contents.content)
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        return wb, contents.sha, None
    except Exception as e:
        return None, None, str(e)

def push_excel_to_github(wb, sha, commit_msg):
    try:
        token = st.secrets.get("GITHUB_TOKEN", None)
        if token is None:
            try: token = st.secrets["github"]["GITHUB_TOKEN"]
            except: pass
        if token is None:
            raise Exception("GITHUB_TOKEN not found in secrets")
        gh = Github(token)
        repo = gh.get_repo(GITHUB_REPO)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        new_content = base64.b64encode(buf.read()).decode("utf-8")
        repo.update_file(EXCEL_FILE_PATH, commit_msg, base64.b64decode(new_content), sha)
        return True, None
    except Exception as e:
        return False, str(e)

INVALID_IMPORT_VALUES = {"", "-", "—", "N/A", "NA", "NONE", "NULL", "VTA", "TBD", "UNKNOWN", "NOT APPLICABLE"}

IMPORT_ALIASES = {
    "seawater": "sea water",
    "sea water": "sea water",
    "phophoric acid": "phosphoric acid",
    "phosporic acid": "phosphoric acid",
    "phosphoricacid": "phosphoric acid",
    "316 ss": "316 stainless steel",
    "316l ss": "316l stainless steel",
    "904l ss": "904l stainless steel",
    "ss316": "316 stainless steel",
    "ss316l": "316l stainless steel",
    "ss904l": "904l stainless steel",
    "pt": "platinum",
}

def _clean_import_value(value):
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    text = re.sub(r"\s+", " ", text)
    if text.upper() in INVALID_IMPORT_VALUES:
        return ""
    return text

def _norm_import_name(value):
    text = _clean_import_value(value).lower()
    if not text:
        return ""
    text = text.replace("&", " and ")
    text = re.sub(r"\bphophoric\b|\bphosporic\b", "phosphoric", text)
    text = re.sub(r"\bstainless\s+steel\b", "ss", text)
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    text = re.sub(r"\s+", " ", text)
    compact = text.replace(" ", "")
    return IMPORT_ALIASES.get(text, IMPORT_ALIASES.get(compact, compact))

def _canonical_map(values):
    result = {}
    for value in values:
        cleaned = _clean_import_value(value)
        norm = _norm_import_name(cleaned)
        if cleaned and norm and norm not in result:
            result[norm] = cleaned
    return result

def _canonical_match(value, candidates, threshold=0.90):
    cleaned = _clean_import_value(value)
    norm = _norm_import_name(cleaned)
    if not cleaned or not norm:
        return "", "empty", 0
    if norm in candidates:
        return candidates[norm], "exact", 100
    best_key, best_score = "", 0
    for key in candidates:
        score = SequenceMatcher(None, norm, key).ratio()
        if score > best_score:
            best_key, best_score = key, score
    if best_key and best_score >= threshold:
        return candidates[best_key], "fuzzy", round(best_score * 100)
    return cleaned, "new", round(best_score * 100)

def get_existing_materials(wb):
    existing = {
        "electrodes": {},
        "liners": {},
        "fluids": {},
        "fluid_materials": {},
        "projects": set(),
    }
    try:
        if "Electrode Materials" in wb.sheetnames:
            ws = wb["Electrode Materials"]
            existing["electrodes"] = _canonical_map(row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row and row[0])
        if "Liner Materials" in wb.sheetnames:
            ws = wb["Liner Materials"]
            existing["liners"] = _canonical_map(row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row and row[0])
        if "Fluid-Material Matrix" in wb.sheetnames:
            ws = wb["Fluid-Material Matrix"]
            fluid_names = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                fluid = _clean_import_value(row[0] if row else "")
                if not fluid:
                    continue
                fluid_names.append(fluid)
                existing["fluid_materials"][_norm_import_name(fluid)] = {
                    "fluid": fluid,
                    "type": _clean_import_value(row[1] if len(row) > 1 else ""),
                    "liner": _clean_import_value(row[4] if len(row) > 4 else ""),
                    "electrode": _clean_import_value(row[5] if len(row) > 5 else ""),
                }
            existing["fluids"] = _canonical_map(fluid_names)
        if "JESA Project Data" in wb.sheetnames:
            ws = wb["JESA Project Data"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    existing["projects"].add(str(row[0]).strip())
    except Exception:
        pass
    return existing

def normalize_imported_instruments(instruments, existing):
    normalized = []
    matched_notes = []
    for inst in instruments:
        item = dict(inst)
        for field, group in [("fluid", "fluids"), ("electrode", "electrodes"), ("liner", "liners")]:
            original = _clean_import_value(item.get(field))
            canonical, status, score = _canonical_match(original, existing.get(group, {}))
            item[f"{field}_original"] = original
            item[f"{field}_match_status"] = status
            item[f"{field}_match_score"] = score
            if status in ("exact", "fuzzy") and canonical:
                item[field] = canonical
                if original and original != canonical:
                    matched_notes.append(f"{field.title()} `{original}` matched existing `{canonical}` for tag {item.get('tag','N/A')}.")
            else:
                item[field] = original
        normalized.append(item)
    return normalized, matched_notes

def detect_new_materials(instruments, existing):
    new_findings = []
    seen = set()

    def add_finding(finding):
        key = (finding.get("type"), _norm_import_name(finding.get("value")), _norm_import_name(finding.get("fluid")), finding.get("tag"))
        if key not in seen:
            seen.add(key)
            new_findings.append(finding)

    for inst in instruments:
        electrode = _clean_import_value(inst.get('electrode'))
        liner = _clean_import_value(inst.get('liner'))
        fluid = _clean_import_value(inst.get('fluid'))
        tag = inst.get('tag', '')

        if electrode and inst.get("electrode_match_status") == "new":
            add_finding({
                "type": "electrode",
                "value": electrode,
                "tag": tag,
                "fluid": fluid,
                "context": f"Found in {tag} for {fluid or 'unknown fluid'}",
                "reason": "New candidate. Datasheet does not contain the full material specification needed for automatic database insertion.",
            })
        if liner and inst.get("liner_match_status") == "new":
            add_finding({
                "type": "liner",
                "value": liner,
                "tag": tag,
                "fluid": fluid,
                "context": f"Found in {tag} for {fluid or 'unknown fluid'}",
                "reason": "New candidate. Datasheet does not contain all compatibility/rating fields needed for automatic database insertion.",
            })
        if fluid and inst.get("fluid_match_status") == "new":
            add_finding({
                "type": "fluid",
                "value": fluid,
                "tag": tag,
                "electrode": electrode,
                "liner": liner,
                "context": f"New fluid candidate in {tag}",
                "reason": "Needs engineering review because maintenance plan, CAPEX score, and full material rules are not provided by the datasheet.",
            })
        elif fluid:
            materials = existing.get("fluid_materials", {}).get(_norm_import_name(fluid), {})
            existing_electrode = materials.get("electrode", "")
            existing_liner = materials.get("liner", "")
            if electrode and existing_electrode and _norm_import_name(electrode) != _norm_import_name(existing_electrode):
                add_finding({
                    "type": "fluid_electrode_variant",
                    "value": electrode,
                    "tag": tag,
                    "fluid": fluid,
                    "matched_existing": existing_electrode,
                    "context": f"Different electrode for {fluid} in {tag} (database: {existing_electrode})",
                    "reason": "Possible valid project variant. Review before changing the fluid-material matrix.",
                })
            if liner and existing_liner and _norm_import_name(liner) != _norm_import_name(existing_liner):
                add_finding({
                    "type": "fluid_liner_variant",
                    "value": liner,
                    "tag": tag,
                    "fluid": fluid,
                    "matched_existing": existing_liner,
                    "context": f"Different liner for {fluid} in {tag} (database: {existing_liner})",
                    "reason": "Possible valid project variant. Review before changing the fluid-material matrix.",
                })
    return new_findings

def apply_updates_to_excel(wb, approved_updates, project_name=""):
    """Append AI discoveries to a review queue instead of corrupting curated DB sheets."""
    sheet_name = "AI Import Review Queue"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        ws.append(["Date", "Project", "Tag", "Item Type", "Extracted Value", "Matched Existing",
                   "Fluid", "Electrode", "Liner", "Source", "Status", "Reason / Notes"])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        widths = [14, 26, 18, 22, 24, 24, 24, 22, 22, 20, 22, 70]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    today = datetime.now().strftime('%Y-%m-%d')
    changes = []
    for update in approved_updates:
        ws.append([
            today,
            project_name,
            update.get("tag", ""),
            update.get("type", ""),
            update.get("value", ""),
            update.get("matched_existing", ""),
            update.get("fluid", ""),
            update.get("electrode", ""),
            update.get("liner", ""),
            "Project Import PDF",
            "Needs I&C Review",
            update.get("reason", update.get("context", "")),
        ])
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        changes.append(f"Queued for review: {update.get('type')} — {update.get('value')}")
    return wb, changes

def save_project_to_gsheet(instruments, project_name):
    try:
        creds_dict = dict(st.secrets["gcp_service_account"])
        creds = Credentials.from_service_account_info(creds_dict, scopes=["https://www.googleapis.com/auth/spreadsheets"])
        client_gs = gspread.authorize(creds)
        spreadsheet = client_gs.open_by_key(SHEET_ID)
        try:
            ws = spreadsheet.worksheet("Project Data")
        except:
            ws = spreadsheet.add_worksheet(title="Project Data", rows=1000, cols=20)
            ws.append_row(["Project","Tag","Service","Fluid","DN","Flow Normal","Flow Max",
                           "Temp Design","Pressure Design","Conductivity","Electrode","Liner",
                           "Tube","Grounding","Accuracy","Vendor","Model","Import Date"])
        today = datetime.now().strftime('%Y-%m-%d')
        saved = 0
        for inst in instruments:
            ws.append_row([inst.get('project', project_name) or project_name, inst.get('tag',''), inst.get('service',''),
                           inst.get('fluid',''), inst.get('dn',''), inst.get('flow_normal',''), inst.get('flow_max',''),
                           inst.get('temp_design',''), inst.get('pressure_design',''), inst.get('conductivity',''),
                           inst.get('electrode',''), inst.get('liner',''), inst.get('tube',''),
                           inst.get('grounding',''), inst.get('accuracy',''), inst.get('vendor',''),
                           inst.get('model',''), today])
            saved += 1
        return saved, None
    except Exception as e:
        return 0, str(e)

def load_project_data_from_gsheet():
    try:
        creds_dict = dict(st.secrets["gcp_service_account"])
        creds = Credentials.from_service_account_info(creds_dict, scopes=["https://www.googleapis.com/auth/spreadsheets"])
        client_gs = gspread.authorize(creds)
        spreadsheet = client_gs.open_by_key(SHEET_ID)
        ws = spreadsheet.worksheet("Project Data")
        return ws.get_all_records()
    except:
        return []

def _num_value(value):
    try:
        if value is None or value == "":
            return None
        return float(str(value).replace(",", ".").replace(">", "").replace("<", "").strip())
    except Exception:
        return None

def _project_value(record, *keys):
    for key in keys:
        if key in record and record.get(key) not in (None, ""):
            return record.get(key)
    return ""

def _standard_project_record(record, source="Internal Excel"):
    return {
        "project": _project_value(record, "project", "Project"),
        "tag": _project_value(record, "tag", "Tag"),
        "service": _project_value(record, "service", "Service"),
        "fluid": _project_value(record, "fluid", "Fluid"),
        "dn": _project_value(record, "dn", "DN"),
        "flow_normal": _project_value(record, "flow_normal", "Flow Normal"),
        "flow_max": _project_value(record, "flow_max", "Flow Max"),
        "temp_design": _project_value(record, "temp_design", "Temp Design"),
        "pressure_design": _project_value(record, "pressure_design", "Pressure Design"),
        "electrode": _project_value(record, "electrode", "Electrode"),
        "liner": _project_value(record, "liner", "Liner"),
        "vendor": _project_value(record, "vendor", "Vendor"),
        "model": _project_value(record, "model", "Model"),
        "import_date": _project_value(record, "Import Date"),
        "source": source,
    }

def get_combined_project_records():
    records = [_standard_project_record(p, "Internal Excel") for p in ai.data.project_data]
    records.extend(_standard_project_record(p, "Imported Project") for p in load_project_data_from_gsheet())
    return [r for r in records if r.get("project") or r.get("tag") or r.get("fluid")]

def compare_imported_projects(fluid, selected_materials, reco_meta, limit=5):
    matches = []
    imported_records = [_standard_project_record(p, "Imported Project") for p in load_project_data_from_gsheet()]
    for p in imported_records:
        score = 0
        details = []
        p_fluid = _clean_import_value(p.get("fluid"))
        if _norm_import_name(p_fluid) and _norm_import_name(p_fluid) == _norm_import_name(fluid):
            score += 35
            details.append(("Fluid", f"{p_fluid} matches selected {fluid}"))

        p_electrode = _clean_import_value(p.get("electrode"))
        if p_electrode and _norm_import_name(p_electrode) == _norm_import_name(getattr(selected_materials, "electrode", "")):
            score += 18
            details.append(("Electrode", f"{p_electrode} matches recommendation"))

        p_liner = _clean_import_value(p.get("liner"))
        if p_liner and _norm_import_name(p_liner) == _norm_import_name(getattr(selected_materials, "liner", "")):
            score += 18
            details.append(("Liner", f"{p_liner} matches recommendation"))

        for label, p_key, meta_key, pts in [
            ("DN", "dn", "dn", 12),
            ("Normal flow", "flow_normal", "flow_normal", 7),
            ("Max flow", "flow_max", "flow_max", 5),
            ("Design temperature", "temp_design", "temp_design", 3),
            ("Design pressure", "pressure_design", "pressure_design", 2),
        ]:
            pv = _num_value(p.get(p_key))
            mv = _num_value(reco_meta.get(meta_key))
            if pv is None or mv is None:
                continue
            tolerance = max(abs(mv) * 0.15, 1.0)
            if abs(pv - mv) <= tolerance:
                score += pts
                details.append((label, f"{pv:g} close to {mv:g}"))

        if score >= 35:
            matches.append({"record": p, "confidence": min(score, 100), "details": details})
    return sorted(matches, key=lambda x: x["confidence"], reverse=True)[:limit]

def generate_maintenance_excel(tag, fluid, cat, mat, vendor_e, vendor_eh, vendor_k, tco, history_records=None, checklist_data=None):
    wb = openpyxl.Workbook()
    vn = vendor_e.model if vendor_e else (vendor_eh.model if vendor_eh else (vendor_k.model if vendor_k else 'N/A'))
    vd = vendor_e.diagnostics if vendor_e else (vendor_eh.diagnostics if vendor_eh else (vendor_k.diagnostics if vendor_k else 'N/A'))
    today = datetime.now().strftime('%Y-%m-%d')
    year = datetime.now().year
    mnt = tco.maintenance
    checklist_data = checklist_data or {}
    def _pf(period, task, col_label):
        """Return saved value for this cell, or None."""
        return checklist_data.get(f"{str(tag).strip()}::{period}::{str(task).strip()}::{col_label}")
    title_font = Font(bold=True, size=14, color="1B2A4A")
    h_font = Font(bold=True, size=11, color="FFFFFF")
    n_font = Font(size=10)
    h_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    g_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    b_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    o_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    p_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
    y_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    r_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    q_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wrap = Alignment(wrap_text=True, vertical='top')
    ws1 = wb.active; ws1.title = "Guideline"
    ws1.column_dimensions['A'].width = 30; ws1.column_dimensions['B'].width = 55
    ws1.merge_cells('A1:B1'); ws1['A1'] = "🔧 MagFlow AI — Instrument Guideline"; ws1['A1'].font = title_font
    ws1['A2'] = f"Generated: {today}"; ws1['A2'].font = Font(italic=True, color="888888")
    row = 4
    ws1.merge_cells(f'A{row}:B{row}'); ws1[f'A{row}'] = "INSTRUMENT INFORMATION"; ws1[f'A{row}'].font = h_font; ws1[f'A{row}'].fill = h_fill
    row += 1
    info = [("Tag Number", tag), ("Service Fluid", fluid), ("Fluid Category", cat),
        ("Electrode", f"{mat.electrode} (Cost: {mat.electrode_cost})"), ("Liner", f"{mat.liner} (Cost: {mat.liner_cost})"),
        ("Body Material", "SS 304L (coil housing)"), ("Tube Material", "SS 316L"),
        ("Grounding", mat.grounding), ("Penetrant Ring", mat.penetrant or "N/A"), ("O-Ring", mat.o_ring),
        ("Flange Coating", mat.flange_coat), ("Vendor Model", vn), ("Diagnostics", vd[:80]),
        ("CAPEX Score", f"{tco.capex_score}/30"), ("Calibration", f"Every {tco.calib_months} months"),
        ("Liner Life", tco.liner_life), ("Electrode Life", tco.electrode_life)]
    for k, v in info:
        ws1[f'A{row}'] = k; ws1[f'A{row}'].font = Font(bold=True); ws1[f'A{row}'].fill = b_fill; ws1[f'A{row}'].border = thin
        ws1[f'B{row}'] = v; ws1[f'B{row}'].border = thin; row += 1
    row += 1
    ws1.merge_cells(f'A{row}:B{row}'); ws1[f'A{row}'] = "DRIFT RISK ASSESSMENT"; ws1[f'A{row}'].font = h_font; ws1[f'A{row}'].fill = h_fill
    row += 1
    for risk in tco.drift_risks:
        icon = "🔴" if risk.level == "High" else "🟡" if risk.level == "Medium" else "🟢"
        fill = r_fill if risk.level == "High" else y_fill if risk.level == "Medium" else g_fill
        ws1.merge_cells(f'A{row}:B{row}')
        ws1[f'A{row}'] = f"{icon} {risk.indicator} — {risk.level}"; ws1[f'A{row}'].font = Font(bold=True, size=11); ws1[f'A{row}'].fill = fill; row += 1
        ws1[f'A{row}'] = "Description:"; ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'B{row}'] = risk.description; ws1[f'B{row}'].alignment = wrap; row += 1
        ws1[f'A{row}'] = "Diagnostic Steps:"; ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'B{row}'] = "\n".join(risk.steps); ws1[f'B{row}'].alignment = wrap
        ws1.row_dimensions[row].height = max(15 * len(risk.steps), 30); row += 1
        ws1[f'A{row}'] = "Frequency:"; ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'B{row}'] = risk.frequency; row += 2
    ws2 = wb.create_sheet("Maintenance Checklist")
    widths = {'A':6,'B':45,'C':9,'D':9,'E':9,'F':9,'G':9,'H':9,'I':9,'J':12,'K':22,'L':20,'M':15,'N':40,'O':24}
    for col, width in widths.items():
        ws2.column_dimensions[col].width = width
    ws2.freeze_panes = "A4"
    ws2.merge_cells('A1:O1'); ws2['A1'] = f"🔧 Maintenance Checklist — {tag}"; ws2['A1'].font = title_font
    ws2['A2'] = f"Fluid: {fluid} | Category: {cat} | Year: {year}"; ws2['A2'].font = Font(italic=True, color="666666")

    done_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws2.add_data_validation(done_dv)
    day_dv = DataValidation(type="list", formula1='"Y"', allow_blank=True)
    ws2.add_data_validation(day_dv)

    def _result_options(task, section):
        text = f"{task} {section}".lower()
        if any(k in text for k in ["alarm", "display", "self-test", "self test", "transmitter"]):
            return '"OK,Alarm found,Reset done,Action required,NC"'
        if any(k in text for k in ["ground", "impedance", "wiring", "connection", "cable", "gland"]):
            return '"OK,<1 Ohm,Loose connection,Moisture found,Action required,NC"'
        if any(k in text for k in ["liner", "electrode", "o-ring", "oring", "ring", "replacement"]):
            return '"OK,Worn,Cleaned,Replaced,Inspection required,NC"'
        if any(k in text for k in ["calibration", "verification", "zero", "re-zero", "bench"]):
            return '"OK,Within tolerance,Adjusted,Calibration required,NC"'
        if any(k in text for k in ["moisture", "ip", "junction"]):
            return '"OK,Dry,Moisture found,Seal repaired,NC"'
        return '"OK,Conform,Observation,Action required,NC"'

    def _add_result_validation(cell, task, section):
        dv = DataValidation(type="list", formula1=_result_options(task, section), allow_blank=True)
        ws2.add_data_validation(dv)
        dv.add(cell)

    def _task_row(row_num, idx, task, period, section, include_days=False, fill=None, compact=False, simple=False):
        ws2.cell(row=row_num, column=1, value=idx).alignment = center
        if compact or simple:
            ws2.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=9)
        ws2.cell(row=row_num, column=2, value=task).font = n_font
        ws2.cell(row=row_num, column=2).alignment = wrap
        if fill:
            ws2.cell(row=row_num, column=1).fill = fill
        for col in range(1, 16):
            c = ws2.cell(row=row_num, column=col)
            c.border = thin
            if col != 2:
                c.alignment = center if col != 14 else wrap
        done_saved = _pf(period, task, "Done?")
        result_saved = _pf(period, task, "Result") or _pf(period, task, "Status")
        tech_saved = _pf(period, task, "Technician")
        date_saved = _pf(period, task, "Date")
        work_saved = _pf(period, task, "Work done")
        if include_days:
            for d, day in enumerate(["Mon","Tue","Wed","Thu","Fri","Sat","Sun"], 3):
                _sv = _pf(period, task, day)
                if _sv is not None:
                    ws2.cell(row=row_num, column=d, value=_sv)
                day_dv.add(ws2.cell(row=row_num, column=d))
        if simple:
            notes_saved = _pf(period, task, "Work done") or _pf(period, task, "Notes")
            if done_saved is not None:
                ws2.cell(row=row_num, column=10, value=done_saved)
            if notes_saved is not None:
                ws2.cell(row=row_num, column=11, value=notes_saved)
            ws2.merge_cells(start_row=row_num, start_column=11, end_row=row_num, end_column=14)
            done_dv.add(ws2.cell(row=row_num, column=10))
            ws2.cell(row=row_num, column=15, value=f'=IF(J{row_num}="Yes","Complete","Not done")')
            return row_num + 1
        if done_saved is not None:
            ws2.cell(row=row_num, column=10, value=done_saved)
        if result_saved is not None:
            ws2.cell(row=row_num, column=11, value=result_saved)
        if tech_saved is not None:
            ws2.cell(row=row_num, column=12, value=tech_saved)
        if date_saved is not None:
            ws2.cell(row=row_num, column=13, value=date_saved)
        if work_saved is not None:
            ws2.cell(row=row_num, column=14, value=work_saved)
        done_dv.add(ws2.cell(row=row_num, column=10))
        _add_result_validation(ws2.cell(row=row_num, column=11), task, section)
        ws2.cell(row=row_num, column=15, value=f'=IF(J{row_num}="Yes",IF(K{row_num}<>"","Complete","Missing result"),"Not done")')
        return row_num + 1

    def _section_summary(row_num, label, start_row, end_row, fill):
        ws2.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=14)
        ws2.cell(row=row_num, column=1, value=f"Validation — {label}").font = Font(bold=True, color="1B2A4A")
        ws2.cell(row=row_num, column=1).fill = fill
        for col in range(1, 16):
            ws2.cell(row=row_num, column=col).border = thin
            ws2.cell(row=row_num, column=col).alignment = center
            ws2.cell(row=row_num, column=col).fill = fill
        ws2.cell(row=row_num, column=15, value=f'=IF(COUNTIF(J{start_row}:J{end_row},"Yes")=ROWS(J{start_row}:J{end_row}),"All tasks done","Pending tasks")')
        ws2.cell(row=row_num, column=15).font = Font(bold=True)
        return row_num + 2

    # ---- Instructions table on the right side (cols Q:S) ----
    ws2.column_dimensions['Q'].width = 20
    ws2.column_dimensions['R'].width = 42
    ws2.column_dimensions['S'].width = 34
    instr_title_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    instr_hdr_fill = PatternFill(start_color="D9ECFF", end_color="D9ECFF", fill_type="solid")
    instr_body_fill = PatternFill(start_color="F8FBFF", end_color="F8FBFF", fill_type="solid")
    instr_alert_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    instr_border = Border(left=Side(style='thin', color="B7C7D9"), right=Side(style='thin', color="B7C7D9"),
                          top=Side(style='thin', color="B7C7D9"), bottom=Side(style='thin', color="B7C7D9"))
    ws2.merge_cells('Q2:S2')
    ws2['Q2'] = "📌 How to Use This Checklist"; ws2['Q2'].font = Font(bold=True, size=13, color="FFFFFF")
    ws2['Q2'].fill = instr_title_fill; ws2['Q2'].alignment = center
    irow = 4
    instr_rows = [
        ("Field", "What to enter", "Notes", True),
        ("No.", "Read-only task number.", "Visual reference only.", False),
        ("Done?", "Select Yes only after doing the task.", "If No/blank, Result can stay empty.", False),
        ("Result", "Pick the result from the dropdown.", "Options adapt to the task type.", False),
        ("Mon → Sun", "For continuous checks, select Y on checked days.", "Blank means not checked that day.", False),
        ("Technician", "Enter initials or full name.", "Example: MH.", False),
        ("Date", "Use YYYY-MM-DD for dated tasks.", "Continuous tasks import as period, e.g. 2026-01 W1.", False),
        ("Work done / Notes", "Describe action, anomaly, or follow-up.", "Keep it short and factual.", False),
        ("Validation", "Auto-calculated.", "Complete / Missing result / Not done.", False),
        ("Cadence", "Filling frequency", "When to use it", True),
        ("Continuous", "Every week.", "Self-test, empty-pipe detection, diagnostics.", False),
        ("Monthly", "Once per month.", "Alarms, display, routine review.", False),
        ("Quarterly", "Every 3 months.", "Inspection, grounding, verification.", False),
        ("Semi-annual", "Every 6 months.", "Liner/electrode checks when applicable.", False),
        ("Annual / Multi-year", "Once a year or every 3-5 years.", "Only Done? and Notes are required.", False),
        ("NC / Action", "If Non-conform, log details.", "Open corrective action and add notes.", "alert"),
    ]
    for label, action, note, row_type in instr_rows:
        vals = [label, action, note]
        ws2.row_dimensions[irow].height = 28 if row_type is True else 34
        for offset, val in enumerate(vals, 17):
            c = ws2.cell(row=irow, column=offset, value=val)
            c.border = instr_border
            c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
            if row_type is True:
                c.font = Font(bold=True, size=10, color="0D47A1")
                c.fill = instr_hdr_fill
            elif row_type == "alert":
                c.font = Font(bold=True, size=10, color="9A3412")
                c.fill = instr_alert_fill
            else:
                c.font = Font(bold=(offset == 17), size=10, color="1F2937")
                c.fill = instr_body_fill
        irow += 1

    row = 4
    months_data = [("January",4),("February",4),("March",4),("April",4),("May",4),("June",4),
                   ("July",4),("August",4),("September",4),("October",4),("November",4),("December",4)]
    quarter_months = {3:"Q1",6:"Q2",9:"Q3",12:"Q4"}
    semester_months = {6:"S1",12:"S2"}
    for m_idx, (month_name, weeks) in enumerate(months_data, 1):
        ws2.merge_cells(f'A{row}:O{row}')
        ws2[f'A{row}'] = f"📅 {month_name} {year}"; ws2[f'A{row}'].font = Font(bold=True, size=12, color="1565C0"); ws2[f'A{row}'].fill = b_fill; row += 1
        if mnt.continuous:
            for w in range(1, weeks + 1):
                ws2.merge_cells(f'A{row}:B{row}')
                ws2[f'A{row}'] = f"   Week {w}"; ws2[f'A{row}'].font = Font(bold=True, size=10, color="2E7D32")
                for d, day in enumerate(["Mon","Tue","Wed","Thu","Fri","Sat","Sun"], 3):
                    c = ws2.cell(row=row, column=d, value=day)
                    c.font = Font(bold=True, size=9); c.alignment = center; c.fill = g_fill; c.border = thin
                for col, label in [(10,"Done?"),(11,"Result"),(12,"Technician"),(13,"Date"),(14,"Work done / Notes"),(15,"Validation")]:
                    c = ws2.cell(row=row, column=col, value=label)
                    c.font = Font(bold=True, size=9); c.alignment = center; c.fill = g_fill; c.border = thin
                row += 1
                start_task = row
                for idx, item in enumerate(mnt.continuous, 1):
                    row = _task_row(row, idx, item, f"{month_name}/Week {w}", "continuous", include_days=True, fill=g_fill)
                row = _section_summary(row, f"{month_name} Week {w}", start_task, row - 1, g_fill)
        if mnt.monthly:
            ws2.merge_cells(f'A{row}:I{row}')
            ws2[f'A{row}'] = f"   📋 Monthly tasks — {month_name}"; ws2[f'A{row}'].font = Font(bold=True, size=10, color="1565C0"); ws2[f'A{row}'].fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
            for col, label in [(10,"Done?"),(11,"Result"),(12,"Technician"),(13,"Date"),(14,"Work done"),(15,"Validation")]:
                ws2.cell(row=row, column=col, value=label).font = Font(bold=True, size=9); ws2.cell(row=row, column=col).fill = b_fill; ws2.cell(row=row, column=col).border = thin
            row += 1
            start_task = row
            for idx, item in enumerate(mnt.monthly, 1):
                row = _task_row(row, idx, item, f"{month_name}/Monthly", "monthly", fill=b_fill, compact=True)
            row = _section_summary(row, f"{month_name} Monthly", start_task, row - 1, b_fill)
        if m_idx in quarter_months and mnt.quarterly:
            ws2.merge_cells(f'A{row}:I{row}')
            ws2[f'A{row}'] = f"   🔍 Quarterly — {quarter_months[m_idx]}"; ws2[f'A{row}'].font = Font(bold=True, size=10, color="E65100"); ws2[f'A{row}'].fill = o_fill
            for col, label in [(10,"Done?"),(11,"Result"),(12,"Technician"),(13,"Date"),(14,"Work done"),(15,"Validation")]:
                ws2.cell(row=row, column=col, value=label).font = Font(bold=True, size=9); ws2.cell(row=row, column=col).fill = o_fill; ws2.cell(row=row, column=col).border = thin
            row += 1
            start_task = row
            for idx, item in enumerate(mnt.quarterly, 1):
                row = _task_row(row, idx, item, f"{month_name}/Quarterly {quarter_months[m_idx]}", "quarterly", fill=o_fill, compact=True)
            row = _section_summary(row, f"{month_name} Quarterly {quarter_months[m_idx]}", start_task, row - 1, o_fill)
        if m_idx in semester_months and mnt.semi_annual:
            ws2.merge_cells(f'A{row}:I{row}')
            ws2[f'A{row}'] = f"   ⚙️ Semi-annual — {semester_months[m_idx]}"; ws2[f'A{row}'].font = Font(bold=True, size=10, color="BF360C"); ws2[f'A{row}'].fill = q_fill
            for col, label in [(10,"Done?"),(11,"Result"),(12,"Technician"),(13,"Date"),(14,"Work done"),(15,"Validation")]:
                ws2.cell(row=row, column=col, value=label).font = Font(bold=True, size=9); ws2.cell(row=row, column=col).fill = q_fill; ws2.cell(row=row, column=col).border = thin
            row += 1
            start_task = row
            for idx, item in enumerate(mnt.semi_annual, 1):
                row = _task_row(row, idx, item, f"{month_name}/Semi-annual {semester_months[m_idx]}", "semi-annual", fill=q_fill, compact=True)
            row = _section_summary(row, f"{month_name} Semi-annual {semester_months[m_idx]}", start_task, row - 1, q_fill)
    if mnt.annual:
        ws2.merge_cells(f'A{row}:O{row}')
        ws2[f'A{row}'] = f"🛠️ ANNUAL MAINTENANCE — {year}"; ws2[f'A{row}'].font = Font(bold=True, size=12, color="6A1B9A"); ws2[f'A{row}'].fill = p_fill; row += 1
        for col, label in [(10,"Done?"),(11,"Notes"),(15,"Validation")]:
            ws2.cell(row=row, column=col, value=label).font = Font(bold=True, size=9); ws2.cell(row=row, column=col).fill = p_fill; ws2.cell(row=row, column=col).border = thin
        ws2.merge_cells(start_row=row, start_column=11, end_row=row, end_column=14)
        row += 1
        start_task = row
        for idx, item in enumerate(mnt.annual, 1):
            row = _task_row(row, idx, item, "ANNUAL", "annual", fill=p_fill, simple=True)
        row = _section_summary(row, "Annual Maintenance", start_task, row - 1, p_fill)
    if mnt.multi_year:
        ws2.merge_cells(f'A{row}:O{row}')
        ws2[f'A{row}'] = "📊 MULTI-YEAR MAINTENANCE (3-5 years)"; ws2[f'A{row}'].font = Font(bold=True, size=12, color="F57F17"); ws2[f'A{row}'].fill = y_fill; row += 1
        for col, label in [(10,"Done?"),(11,"Notes"),(15,"Validation")]:
            ws2.cell(row=row, column=col, value=label).font = Font(bold=True, size=9); ws2.cell(row=row, column=col).fill = y_fill; ws2.cell(row=row, column=col).border = thin
        ws2.merge_cells(start_row=row, start_column=11, end_row=row, end_column=14)
        row += 1
        start_task = row
        for idx, item in enumerate(mnt.multi_year, 1):
            row = _task_row(row, idx, item, "MULTI-YEAR", "multi-year", fill=y_fill, simple=True)
        row = _section_summary(row, "Multi-year Maintenance", start_task, row - 1, y_fill)
    if mnt.replacement:
        ws2.merge_cells(f'A{row}:O{row}')
        ws2[f'A{row}'] = "♻️ COMPONENT REPLACEMENT"; ws2[f'A{row}'].font = Font(bold=True, size=12, color="C62828"); ws2[f'A{row}'].fill = r_fill; row += 1
        for col, label in [(10,"Done?"),(11,"Notes"),(15,"Validation")]:
            ws2.cell(row=row, column=col, value=label).font = Font(bold=True, size=9); ws2.cell(row=row, column=col).fill = r_fill; ws2.cell(row=row, column=col).border = thin
        ws2.merge_cells(start_row=row, start_column=11, end_row=row, end_column=14)
        row += 1
        start_task = row
        for idx, item in enumerate(mnt.replacement, 1):
            row = _task_row(row, idx, item, "REPLACEMENT", "replacement", fill=r_fill, simple=True)
        row = _section_summary(row, "Component Replacement", start_task, row - 1, r_fill)
    ws3 = wb.create_sheet("Historical Data")
    for col, w in [(1,15),(2,20),(3,20),(4,35),(5,15),(6,20),(7,30)]:
        ws3.column_dimensions[get_column_letter(col)].width = w
    ws3.merge_cells('A1:G1')
    ws3['A1'] = f"📋 Maintenance History — {tag}"; ws3['A1'].font = title_font
    headers = ["Date", "Tag Number", "Type", "Task Performed", "Result", "Technician", "Notes"]
    for col, h in enumerate(headers, 1):
        c = ws3.cell(row=3, column=col, value=h)
        c.font = h_font; c.fill = h_fill; c.border = thin; c.alignment = center
    if history_records:
        for r_idx, record in enumerate(history_records, 4):
            for col, val in enumerate([record.get('Date',''), record.get('Tag Number',''), record.get('Type',''),
                                        record.get('Task',''), record.get('Result',''), record.get('Technician',''), record.get('Notes','')], 1):
                c = ws3.cell(row=r_idx, column=col, value=val)
                c.border = thin
                c.fill = b_fill if r_idx % 2 == 0 else PatternFill(fill_type=None)
    else:
        ws3.merge_cells('A4:G4')
        ws3['A4'] = "No interventions recorded yet. Use the 'Maintenance History' tab in MagFlow AI to log interventions."
        ws3['A4'].font = Font(italic=True, color="888888")
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ============================================================
#  PAGE ROUTER  —  landing page is NOT a tab; it's shown first
# ============================================================
if "page" not in st.session_state:
    st.session_state.page = "home"

# ------------------------------------------------------------
#  LANDING PAGE
# ------------------------------------------------------------
def render_landing():
    logo_html = f"<img class='mf-landing-logo' src='{IMG_JESA_LOGO}' alt='JESA logo'>" if IMG_JESA_LOGO and not IMG_JESA_LOGO.startswith("REPLACE_") else ""
    st.markdown(
        f"""
        <div class="mf-landing-head">
          {logo_html}
          <div>
            <div class="mf-landing-title">MagFlow <span class="ai">AI</span></div>
            <div class="mf-landing-sub">AI-Based Predictive Maintenance System for Electromagnetic Flowmeters</div>
            <div class="mf-landing-meta">PFE ENSA — JESA (OCP × Worley) · Instrumentation &amp; Control · 2026</div>
          </div>
        </div>
        """,
        unsafe_allow_html=True)

    # Hero image
    if IMG_HERO and not IMG_HERO.startswith("REPLACE_"):
        st.markdown(f"<div class='mf-hero'><img src='{IMG_HERO}' alt='JESA industrial hero'></div>",
                    unsafe_allow_html=True)
    else:
        st.info("🖼️ Hero image goes here — set `IMG_HERO` at the top of app.py (see chat for how to get the URL).")

    st.markdown("<br>", unsafe_allow_html=True)

    # Enter button (top, easy to find)
    ec1, ec2, ec3 = st.columns([1, 2, 1])
    with ec2:
        if st.button("🚀 Enter MagFlow AI", type="primary", use_container_width=True):
            st.session_state.page = "app"
            st.rerun()

    # What it is / what it does
    st.markdown("<div class='mf-home-title'>What is MagFlow AI?</div>", unsafe_allow_html=True)
    st.markdown(
        "<p class='mf-home-copy'>MagFlow AI is an AI-based predictive-maintenance assistant "
        "<b>developed and validated on electromagnetic flowmeters, with a generic and extensible architecture</b>. "
        "It turns a flowmeter's process conditions into a complete engineering recommendation: the right materials, "
        "compatible vendor models, total cost of ownership, drift risk, and a year-round maintenance plan — "
        "grounded in a database of real JESA/OCP projects.</p>",
        unsafe_allow_html=True)

    st.markdown("<div class='mf-home-subtitle'>What it does</div>", unsafe_allow_html=True)
    f1, f2, f3, f4 = st.columns(4)
    for col, (icon, title, body, color) in zip(
        [f1, f2, f3, f4],
        [("✅", "Validation", "Checks if a magnetic flowmeter suits your fluid, velocity and conductivity.", "#2E7D32"),
         ("🧪", "Materials", "Selects electrode, liner, grounding, O-ring and coating for the fluid.", "#1565C0"),
         ("🏭", "Vendors", "Matches compatible models from Emerson, Endress+Hauser, Krohne & others.", "#E65100"),
         ("📊", "TCO & Drift", "Estimates cost of ownership, drift risk and a full maintenance plan.", "#6A1B9A")]):
        with col:
            st.markdown(
                f"<div class='mf-card'><div style='font-size:30px'>{icon}</div>"
                f"<p style='font-size:17px;font-weight:700;color:{color};margin:6px 0'>{title}</p>"
                f"<p style='font-size:14px;color:#555'>{body}</p></div>",
                unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Magnetic flowmeter — photo + description
    st.markdown("<div class='mf-home-subtitle'>The magnetic flowmeter</div>", unsafe_allow_html=True)
    pc1, pc2 = st.columns([2, 3])
    with pc1:
        if IMG_FLOWMETER and not IMG_FLOWMETER.startswith("REPLACE_"):
            st.image(IMG_FLOWMETER, use_container_width=True, caption="Electromagnetic flowmeter")
        else:
            st.info("🖼️ Flowmeter photo goes here — set `IMG_FLOWMETER` at the top of app.py.")
    with pc2:
        st.markdown(
            "<div class='mf-card'>"
            "<p style='font-size:17px;font-weight:700;color:#1B2A4A;margin:0 0 8px 0'>Magnetic Flowmeter</p>"
            "<p style='font-size:15px;color:#444;line-height:1.6;margin:0'>"
            "A magnetic (electromagnetic) flowmeter measures the flow of conductive liquids using "
            "Faraday's law of induction. As the fluid passes through a magnetic field, it generates a "
            "voltage proportional to its velocity, picked up by two electrodes in the tube wall. "
            "With no moving parts and no obstruction in the flow path, it produces no pressure drop and "
            "needs little maintenance — making it ideal for water, slurries, acids, and abrasive fluids "
            "common in OCP/JESA process lines.</p></div>",
            unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # How to use it (first-time user steps)
    st.markdown("<div class='mf-home-title'>How to use it</div>", unsafe_allow_html=True)
    steps = [
        ("1 · Enter process data", "Open the <b>Recommendation Engine</b>, then fill the fluid, pipe, operating conditions and any special notes."),
        ("2 · Run the engine", "Click <b>Run Recommendation</b>. You'll get all four layers: validation, materials, vendors, and TCO & drift."),
        ("3 · Get your maintenance sheet", "Download the maintenance Excel — Guideline, full month-by-month checklist, and historical data."),
        ("4 · Log interventions over time", "Use <b>Maintenance History</b> to record each intervention by tag — years later, search the tag to find your work."),
        ("5 · Import projects", "Use <b>Project Import</b> to upload a datasheet PDF; the AI extracts instruments and flags new materials."),
    ]
    for title, body in steps:
        st.markdown(f"<div class='mf-step'><b>{title}</b><br>{body}</div>", unsafe_allow_html=True)

    st.divider()
    ec1, ec2, ec3 = st.columns([1, 2, 1])
    with ec2:
        if st.button("🚀 Get Started", type="primary", use_container_width=True, key="enter_bottom"):
            st.session_state.page = "app"
            st.rerun()

if st.session_state.page == "home":
    render_landing()
    st.stop()

# ============================================================
#  SIDEBAR (only inside the app, not on landing)
# ============================================================
with st.sidebar:
    if IMG_JESA_LOGO and not IMG_JESA_LOGO.startswith("REPLACE_"):
        st.image(IMG_JESA_LOGO, width=155)
    if st.button("🏠  Home", use_container_width=True):
        st.session_state.page = "home"
        st.rerun()
    st.divider()
    st.markdown("### 🤖 MagFlow Assistant")
    lang_options = {"🇫🇷 French": "fr-FR", "🇬🇧 English": "en-US", "🇸🇦 Arabic": "ar-SA", "🇪🇸 Spanish": "es-ES"}
    selected_lang = st.selectbox("Voice language", list(lang_options.keys()), label_visibility="collapsed")
    lang_code = lang_options[selected_lang]
    voice_html = f"""<div style="display:flex;gap:8px;align-items:center;margin-top:12px">
    <button id="voiceBtn" onclick="startVoice()" style="background:#1B2A4A;color:white;border:none;padding:10px 18px;border-radius:10px;cursor:pointer;font-size:14px;font-weight:700">🎤 Tap to speak</button>
    <button id="sendBtn" onclick="sendVoice()" style="background:#0B55D9;color:white;border:none;padding:10px 18px;border-radius:10px;cursor:pointer;font-size:14px;font-weight:700;display:none">📤 Send</button>
    <span id="status" style="font-size:12px;color:#667085"></span></div>
    <script>
    let recognition=null,transcript='';
    function startVoice(){{
        if(!('webkitSpeechRecognition'in window||'SpeechRecognition'in window)){{document.getElementById('status').innerText='Chrome only';return;}}
        const SR=window.SpeechRecognition||window.webkitSpeechRecognition;
        recognition=new SR();recognition.lang='{lang_code}';recognition.continuous=true;recognition.interimResults=true;
        recognition.onstart=()=>{{document.getElementById('voiceBtn').innerText='🔴 Listening...';document.getElementById('status').innerText='';document.getElementById('sendBtn').style.display='inline-block';}};
        recognition.onresult=(e)=>{{transcript=Array.from(e.results).map(r=>r[0].transcript).join('');document.getElementById('status').innerText=transcript.slice(0,50)+'...';}};
        recognition.onerror=()=>{{document.getElementById('voiceBtn').innerText='🎤 Tap to speak';document.getElementById('sendBtn').style.display='none';}};
        recognition.start();
    }}
    function sendVoice(){{
        if(recognition)recognition.stop();
        if(transcript)window.parent.postMessage({{type:'voice_text', text:transcript}}, '*');
        document.getElementById('voiceBtn').innerText='🎤 Tap to speak';
        document.getElementById('sendBtn').style.display='none';
        transcript='';
    }}
    </script>"""
    st.components.v1.html(voice_html, height=80)
    vq = st.query_params.get("voice", None)
    if vq:
        st.query_params.clear()
        st.session_state.voice_input = vq
    if "messages" not in st.session_state:
        st.session_state.messages = [{"role": "assistant", "content": "👋 How can I help you today?"}]

    latest_assistant = next((m["content"] for m in reversed(st.session_state.messages) if m["role"] == "assistant"), "👋 How can I help you today?")
    latest_assistant_html = html.escape(latest_assistant)
    st.markdown("<div class='mf-assistant-label'>Or type:</div>", unsafe_allow_html=True)
    st.markdown(
        "<div class='mf-assistant-greeting'><span class='mf-bot-badge'>🤖</span>"
        f"<span>{latest_assistant_html}</span></div>",
        unsafe_allow_html=True
    )

    vt = st.session_state.pop("voice_input", None)
    typed_prompt = st.text_area("Ask a question...", key="assistant_text", height=110, label_visibility="collapsed", placeholder="Ask a question...")
    send_prompt = st.button("Send", key="assistant_send", use_container_width=True)
    prompt = vt or (typed_prompt.strip() if send_prompt and typed_prompt.strip() else None)
    if prompt:
        st.session_state.messages.append({"role": "user", "content": prompt})
        lr = search_local(prompt)
        ch = [{"role": m["role"], "content": m["content"]} for m in st.session_state.messages[1:] if m["role"] in ["user","assistant"]][:-1]
        with st.spinner("Thinking..."):
            response = ask_claude(prompt, lr, ch)
        st.session_state.messages.append({"role": "assistant", "content": response})
        st.markdown(f"<div class='mf-assistant-reply'>{html.escape(response)}</div>", unsafe_allow_html=True)
    if st.button("Clear", use_container_width=True):
        st.session_state.messages = [{"role": "assistant", "content": "👋 Cleared."}]

# ============================================================
#  APP HEADER + TABS  (Dashboard moved to LAST)
# ============================================================
st.markdown("""
<div class="mf-app-hero">
  <div class="mf-hero-inner">
    <div class="mf-wave-logo"></div>
    <div>
      <div class="mf-app-title">MagFlow <span class="ai">AI</span></div>
      <p class="mf-app-sub">AI-Based Predictive Maintenance System for Electromagnetic Flowmeters</p>
      <p class="mf-app-meta">PFE ENSA — JESA (OCP × Worley) &nbsp; | &nbsp; Instrumentation &amp; Control &nbsp; | &nbsp; Developed by Maroua Hakkak — 2026</p>
    </div>
  </div>
 </div>
""", unsafe_allow_html=True)
st.markdown('<div class="mf-home-row">', unsafe_allow_html=True)
_home_gap, _home_btn, _home_gap2 = st.columns([4.5, 1.2, 4.5])
with _home_btn:
    if st.button("🏠 Home", key="top_home_modern", use_container_width=True):
        st.session_state.page = "home"
        st.rerun()
st.markdown('</div>', unsafe_allow_html=True)

# New order: Recommendation Engine → Maintenance History → Project Import → Dashboard
mt_reco, mt_hist, mt_import, mt_dash = st.tabs(
    ["🔧 Recommendation Engine", "📋 Maintenance History", "📂 Project Import", "📊 Dashboard"])

# ------------------------------------------------------------
#  MAINTENANCE HISTORY
# ------------------------------------------------------------
with mt_hist:
    st.header("📋 Maintenance History")
    st.markdown("Log and track all maintenance interventions for each flowmeter across JESA/OCP projects.")
    with st.expander("📥 Import from Maintenance Excel (Historical Data sheet)", expanded=False):
        st.markdown("Upload a filled MagFlow AI maintenance Excel. Enter the **tag** below, then import. "
                    "The app will: (1) save the **Historical Data** rows, (2) auto-create interventions from every "
                    "**filled checklist task**, and (3) remember the "
                    "checklist so it stays pre-filled when you re-download it for the same tag.")
        uploaded_excel = st.file_uploader("Upload Excel file", type=["xlsx"], key="excel_import")
        if uploaded_excel:
            imp_tag = st.text_input("Tag Number for this checklist", placeholder="e.g., 204M-FE/FIT-063M", key="imp_tag")
            if st.button("📤 Import to Database", type="primary"):
                with st.spinner("Reading Excel and saving to database..."):
                    count, error = import_from_excel(uploaded_excel)
                    chk_saved = 0
                    auto_interv = 0
                    detected_interv = 0
                    if imp_tag.strip():
                        # 1) remember the filled checklist cells (for pre-fill on re-download)
                        try:
                            uploaded_excel.seek(0)
                        except Exception:
                            pass
                        cells = read_checklist_from_excel(uploaded_excel, imp_tag.strip())
                        chk_saved = save_checklist_records(imp_tag.strip(), cells)
                        # 2) auto-generate interventions in Historical Data from the checklist
                        try:
                            uploaded_excel.seek(0)
                        except Exception:
                            pass
                        interv = extract_interventions_from_checklist(uploaded_excel, imp_tag.strip())
                        detected_interv = len(interv)
                        auto_interv = save_interventions_dedup(interv)
                if error: st.error(f"❌ Import failed: {error}")
                elif count == 0 and chk_saved == 0 and auto_interv == 0 and detected_interv == 0: st.warning("⚠️ No valid rows found.")
                else:
                    parts = []
                    if count: parts.append(f"{count} row(s) from Historical Data")
                    if auto_interv: parts.append(f"{auto_interv} intervention(s) auto-created from checklist")
                    elif imp_tag.strip() and detected_interv:
                        parts.append(f"{detected_interv} checklist intervention(s) detected but not added (already saved or Google Sheets write failed)")
                    if chk_saved: parts.append(f"{chk_saved} checklist cell(s) saved for pre-fill")
                    st.success("✅ " + " · ".join(parts) if parts else "✅ Done")
                    st.balloons()
    st.divider()
    col_form, col_view = st.columns([1, 1])
    with col_form:
        st.subheader("➕ Log New Intervention")
        with st.form("history_form"):
            h_tag = st.text_input("Tag Number *", placeholder="e.g., 204M-FE/FIT-063M")
            h_date = st.date_input("Date *", value=datetime.now())
            h_type = st.selectbox("Intervention Type *", ["Calibration","Visual Inspection","Liner Inspection","Electrode Check","Grounding Verification","Zero Reset","Component Replacement","Corrective Maintenance","Other"])
            h_task = st.text_area("Task Performed *", placeholder="Describe what was done...", height=100)
            h_result = st.selectbox("Result *", ["Conform","Non-conform","Replaced","Adjusted","Pending"])
            h_tech = st.text_input("Technician Name *", placeholder="Full name")
            h_notes = st.text_area("Notes", placeholder="Additional observations...", height=80)
            submitted = st.form_submit_button("💾 Save Intervention", use_container_width=True, type="primary")
            if submitted:
                if not h_tag or not h_task or not h_tech: st.error("Please fill in Tag Number, Task, and Technician Name.")
                else:
                    success = save_history(str(h_date), h_tag.strip(), h_type, h_task, h_result, h_tech, h_notes)
                    if success: st.success(f"✅ Intervention saved for {h_tag}"); st.balloons()
                    else: st.error("❌ Error saving. Check Google Sheets connection.")
    with col_view:
        st.subheader("🔍 View History by Tag")
        search_tag = st.text_input("Search Tag Number", placeholder="e.g., 204M-FE/FIT-063M")
        if search_tag:
            with st.spinner("Loading history..."):
                records = load_history(tag_filter=search_tag)
            if records:
                st.success(f"Found {len(records)} intervention(s) for **{search_tag}**")
                for r in records:
                    result_text = str(r.get('Result','')).strip()
                    rc = "🔴" if _is_nonconform(result_text) else "🟢" if _is_positive_result(result_text) else "🟡"
                    tech_text = str(r.get('Technician','')).strip()
                    title_bits = [r.get('Date',''), r.get('Type',''), result_text]
                    if tech_text: title_bits.append(f"Tech: {tech_text}")
                    with st.expander(f"{rc} " + " | ".join([str(x) for x in title_bits if x])):
                        st.markdown(f"**Task:** {r.get('Task','')}"); st.markdown(f"**Result:** {r.get('Result','')}"); st.markdown(f"**Technician:** {r.get('Technician','')}")
                        if r.get('Notes'): st.markdown(f"**Notes:** {r.get('Notes','')}")
            else: st.info(f"No history found for tag **{search_tag}**")
        st.divider()
        st.subheader("📊 All Interventions by Tag")
        with st.spinner("Loading..."):
            all_records = load_history()
        if all_records:
            st.caption(f"Total interventions recorded: **{len(all_records)}**")
            tags = {}
            for r in all_records:
                tag = r.get('Tag Number', 'Unknown')
                if tag not in tags: tags[tag] = []
                tags[tag].append(r)
            for tag, records in sorted(tags.items()):
                cc = sum(1 for r in records if _is_positive_result(r.get('Result')))
                nc = sum(1 for r in records if _is_nonconform(r.get('Result')))
                oc = len(records) - cc - nc
                summary = (f"🟢 {cc}" if cc else "") + (f"  🔴 {nc}" if nc else "") + (f"  🟡 {oc}" if oc else "")
                with st.expander(f"📁 **{tag}** — {len(records)} intervention(s)  {summary}"):
                    for r in sorted(records, key=lambda x: x.get('Date',''), reverse=True):
                        result_text = str(r.get('Result','')).strip()
                        rc = "🔴" if _is_nonconform(result_text) else "🟢" if _is_positive_result(result_text) else "🟡"
                        tech_text = str(r.get('Technician','')).strip()
                        st.markdown(f"{rc} **{r.get('Date','')}** | {r.get('Type','')} | **{result_text or 'No result'}**" + (f" | {tech_text}" if tech_text else ""))
                        if r.get('Task'): st.caption(f"↳ {r.get('Task','')[:80]}{'...' if len(r.get('Task','')) > 80 else ''}")
        else: st.info("No interventions recorded yet.")

# ------------------------------------------------------------
#  DASHBOARD  (now the LAST tab)
# ------------------------------------------------------------
with mt_dash:
    st.header("📊 JESA Project Analytics")
    project_data_all = get_combined_project_records()
    imported_count = sum(1 for p in project_data_all if p.get("source") == "Imported Project")
    internal_count = sum(1 for p in project_data_all if p.get("source") == "Internal Excel")
    if imported_count:
        st.caption(f"Showing **{len(project_data_all)}** project instrument records: **{internal_count}** internal Excel + **{imported_count}** imported from Project Import.")
    pc = {}
    for p in project_data_all:
        pn = p.get('project','')
        if pn and not any(k in pn for k in ['LEGEND','NOTE','Original','UPDATED']): pc[pn]=pc.get(pn,0)+1
    for pn,c in sorted(pc.items(), key=lambda x:-x[1]):
        st.markdown(f"**{pn}** — {c}"); st.progress(c/max(pc.values()))
    d1,d2 = st.columns(2)
    with d1:
        st.subheader("Electrodes")
        ec = {}
        for p in project_data_all:
            e = p.get('electrode','').strip()
            if e and e!='N/A': ec[e]=ec.get(e,0)+1
        for en,c in sorted(ec.items(),key=lambda x:-x[1])[:10]:
            st.markdown(f"**{en}** — {c}"); st.progress(c/max(ec.values()))
    with d2:
        st.subheader("Liners")
        lc = {}
        for p in project_data_all:
            l = p.get('liner','').strip()
            if l and l!='N/A': lc[l]=lc.get(l,0)+1
        for ln,c in sorted(lc.items(),key=lambda x:-x[1])[:10]:
            st.markdown(f"**{ln}** — {c}"); st.progress(c/max(lc.values()))
    s1,s2,s3,s4,s5,s6 = st.columns(6)
    with s1: st.metric("Fluids", len(ai.data.fluids))
    with s2: st.metric("Electrodes", len(ai.data.electrodes))
    with s3: st.metric("Liners", len(ai.data.liners))
    with s4: st.metric("Models", len(ai.data.vendor_models))
    with s5: st.metric("Drift", len(ai.data.drift_indicators))
    with s6: st.metric("Projects", len(project_data_all), delta=f"+{imported_count} imported" if imported_count else None)

# ------------------------------------------------------------
#  RECOMMENDATION ENGINE
# ------------------------------------------------------------
with mt_reco:
    tab1,tab2,tab3,tab4 = st.tabs(["🧪 Process Fluid","📐 Pipe Dimensions","🌡️ Operating Conditions","📝 Special Notes"])
    with tab1:
        c1,c2 = st.columns(2)
        with c1:
            fluid = st.selectbox("Service Fluid", ai.get_fluid_names())
            tag_number = st.text_input("🏷️ Tag Number", placeholder="e.g., 204M-FE/FIT-063M")
        with c2:
            cat = ai.get_fluid_category(fluid)
            colors = {'Corrosive':'🔴','Abrasive':'🟠','Charged':'🟡','Clean':'🟢','Unknown':'⚪'}
            st.markdown(f"<p style='font-size:17px;font-weight:600;color:#1B2A4A;margin-bottom:4px'>Fluid Category: {colors.get(cat,'')} {cat}</p>", unsafe_allow_html=True)
            conductivity = st.number_input("Conductivity (µS/cm)", value=5000.0, min_value=0.0, step=100.0)
    with tab2:
        c1,c2,c3 = st.columns(3)
        with c1: dn = st.number_input("DN (mm)", value=80, min_value=3, max_value=3000, step=10)
        with c2: pipe_mat = st.selectbox("Pipe Material", ["FRP","Stainless Steel","Carbon Steel","ZeCor","PVC","HDPE"])
        with c3: pipe_liner = st.selectbox("Pipe Liner", ["Not applicable","HDPE","Rubber lined","Epoxy coated","Glass lined"])
        pipe_thick = st.number_input("Pipe Thickness (mm)", value=0.0, min_value=0.0, step=0.5)
    with tab3:
        st.subheader("Flow")
        fc1,fc2 = st.columns(2)
        with fc1: flow_normal = st.number_input("Normal Flow (m³/h)", value=23.0, min_value=0.0, step=1.0)
        with fc2: flow_max = st.number_input("Max Flow (m³/h)", value=26.0, min_value=0.0, step=1.0)
        st.subheader("Temperature (°C)")
        tc1,tc2,tc3,tc4 = st.columns(4)
        with tc1: t_min = st.number_input("Min Temp", value=0.0, step=1.0)
        with tc2: t_norm = st.number_input("Normal Temp", value=82.0, step=1.0)
        with tc3: t_max = st.number_input("Max Temp", value=95.0, step=1.0)
        with tc4: t_des = st.number_input("Design Temp", value=105.0, step=1.0)
        st.subheader("Pressure (bar)")
        pc1,pc2,pc3,pc4,pc5 = st.columns(5)
        with pc1: p_min = st.number_input("Min P", value=0.0, step=0.5)
        with pc2: p_norm = st.number_input("Normal P", value=5.0, step=0.5)
        with pc3: p_max = st.number_input("Max P", value=8.0, step=0.5)
        with pc4: p_des = st.number_input("Design P", value=10.0, step=0.5)
        with pc5: p_drop = st.number_input("ΔP Allow.", value=0.0, step=0.1)
        st.subheader("Fluid Properties")
        fp1,fp2 = st.columns(2)
        with fp1: visc = st.number_input("Viscosity (cP)", value=16.0, min_value=0.0)
        with fp2: dens = st.number_input("Density (kg/m³)", value=1730.0, min_value=0.0)
    with tab4:
        special = st.text_area("Special Conditions", placeholder="ATEX, SIL 2...", height=80)
        notes = st.text_area("User Notes", placeholder="Prefers Emerson...", height=80)
    st.divider()

    # ---- Run: compute and STORE results in session_state (so downloads/lang don't wipe them) ----
    if st.button("🚀 Run Recommendation", type="primary", use_container_width=True):
        inp = ProcessInput(fluid_name=fluid, pipe_material=pipe_mat, pipe_thickness=pipe_thick,
            pipe_liner=pipe_liner, tube_material='SS 316L', dn=dn, flow_normal=flow_normal, flow_max=flow_max,
            temp_min=t_min, temp_normal=t_norm, temp_max=t_max, temp_design=t_des,
            pressure_min=p_min, pressure_normal=p_norm, pressure_max=p_max, pressure_design=p_des,
            pressure_drop=p_drop, conductivity=conductivity, viscosity=visc, density=dens,
            special_conditions=special, user_notes=notes)
        result = ai.recommend(inp)
        st.session_state["reco_result"] = result
        st.session_state["reco_meta"] = {
            "fluid": fluid, "cat": cat,
            "tag": tag_number if tag_number else "NEW-INSTRUMENT",
            "dn": dn,
            "flow_normal": flow_normal,
            "flow_max": flow_max,
            "temp_design": t_des,
            "pressure_design": p_des,
        }

    # ---- Render results from session_state (survives any re-run) ----
    if "reco_result" in st.session_state:
        result = st.session_state["reco_result"]
        meta = st.session_state["reco_meta"]
        fluid = meta["fluid"]; cat = meta["cat"]
        v,m,vendors,tco = result['validation'],result['materials'],result['vendors'],result['tco']

        with st.container(border=True):
            st.markdown("<div style='background-color:#E8F5E9;padding:15px;border-radius:10px;border-left:5px solid #2E7D32'><h2 style='color:#2E7D32;margin:0'>Layer 1 — Input Validation</h2></div>", unsafe_allow_html=True)
            st.write("")
            if v.is_valid:
                st.success(f"✅ Magflow suitable — Velocity: {v.velocity} m/s — Category: {v.fluid_category}")
                if v.recommended_dn > 0: st.caption(f"💡 DN {v.recommended_dn} mm recommended")
            else:
                st.error("❌ Magflow NOT suitable")
                for e in v.errors: st.error(e)
            for w in v.warnings: st.warning(w)

        if not v.is_valid:
            st.stop()

        with st.container(border=True):
            st.markdown("<div style='background-color:#E3F2FD;padding:15px;border-radius:10px;border-left:5px solid #1565C0'><h2 style='color:#1565C0;margin:0'>Layer 2 — Material Selection</h2></div>", unsafe_allow_html=True)
            st.write("")
            c1,c2,c3,c4 = st.columns(4)
            with c1:
                ab = " ⚠️ SO" if 'special' in m.electrode_avail.lower() else ""
                st.markdown(f"<p style='font-size:18px;font-weight:bold;margin-bottom:2px'>Electrode</p><p style='font-size:15px;color:#333'>{m.electrode}{ab}</p><p style='font-size:12px;color:#2E7D32'>Cost: {m.electrode_cost}</p>", unsafe_allow_html=True)
                if m.electrode_alt: st.caption(f"Alt: {', '.join(m.electrode_alt)}")
            with c2:
                ab2 = " ⚠️ SO" if 'special' in m.liner_avail.lower() else ""
                st.markdown(f"<p style='font-size:18px;font-weight:bold;margin-bottom:2px'>Liner</p><p style='font-size:15px;color:#333'>{m.liner}{ab2}</p><p style='font-size:12px;color:#2E7D32'>Cost: {m.liner_cost}</p>", unsafe_allow_html=True)
                if m.liner_alt: st.caption(f"Alt: {', '.join(m.liner_alt)}")
            with c3: st.markdown(f"<p style='font-size:18px;font-weight:bold;margin-bottom:2px'>Body Material</p><p style='font-size:15px;color:#333'>SS 304L</p><p style='font-size:12px;color:#666'>Coil housing</p>", unsafe_allow_html=True)
            with c4: st.markdown(f"<p style='font-size:18px;font-weight:bold;margin-bottom:2px'>Grounding</p><p style='font-size:15px;color:#333'>{m.grounding}</p>", unsafe_allow_html=True)
            c5,c6,c7 = st.columns(3)
            with c5: st.markdown(f"<p style='font-size:18px;font-weight:bold;margin-bottom:2px'>Penetrant Ring</p><p style='font-size:15px;color:#333'>{m.penetrant or 'N/A'}</p>", unsafe_allow_html=True)
            with c6: st.markdown(f"<p style='font-size:18px;font-weight:bold;margin-bottom:2px'>O-Ring</p><p style='font-size:15px;color:#333'>{m.o_ring}</p>", unsafe_allow_html=True)
            with c7: st.markdown(f"<p style='font-size:18px;font-weight:bold;margin-bottom:2px'>Flange Coating</p><p style='font-size:15px;color:#333'>{m.flange_coat}</p>", unsafe_allow_html=True)
            if m.liner_dn_warning: st.warning(f"📐 {m.liner_dn_warning}")
            if m.tube_warning: st.info(f"🔧 {m.tube_warning}")
            for w in m.warnings: st.warning(w)
            if m.remarks: st.info(f"📝 {m.remarks}")

        with st.container(border=True):
            st.markdown("<div style='background-color:#FFF3E0;padding:15px;border-radius:10px;border-left:5px solid #E65100'><h2 style='color:#E65100;margin:0'>Layer 3 — Vendor Recommendations</h2></div>", unsafe_allow_html=True)
            st.write("")
            st.caption(f"Compatible: {len(vendors['all'])} / {len(ai.data.vendor_models)} models")
            vc1,vc2,vc3 = st.columns(3)
            for cw,label,key,vname in [(vc1,"🔴 EMERSON","emerson","Emerson"),(vc2,"🔵 ENDRESS+HAUSER","eh","Endress+Hauser"),(vc3,"🟢 KROHNE","krohne","Krohne")]:
                with cw:
                    url = vendor_url(vname)
                    head = f"<a href='{url}' target='_blank' style='text-decoration:none;color:#1565C0'>{label} 🔗</a>" if url else label
                    st.markdown(f"<p style='font-size:18px;font-weight:bold'>{head}</p>", unsafe_allow_html=True)
                    rec = vendors[key]
                    if rec:
                        so = " ⚠️ SO" if rec.special_order else ""
                        st.markdown(f"<p style='font-size:16px;font-weight:bold;color:#333'>{rec.model}</p><p style='font-size:12px;color:#666'>{rec.model_type}{so}</p>", unsafe_allow_html=True)
                        st.markdown(f"**Accuracy:** {rec.accuracy}"); st.markdown(f"**Pressure:** {rec.pressure}")
                        st.markdown(f"**DN:** {rec.pipe_sizes}"); st.markdown(f"**Excitation:** {rec.excitation}")
                        st.markdown(f"**IP:** {rec.ip}"); st.markdown(f"**Protocols:** {rec.protocols}"); st.markdown(f"**Diagnostics:** {rec.diagnostics}")
                    else: st.warning("No compatible model")
            with st.expander("🌍 Other Vendors"):
                for ev in vendors.get('extra',[]):
                    url = vendor_url(ev['vendor'])
                    name = f"<a href='{url}' target='_blank'>{ev['vendor']}</a>" if url else ev['vendor']
                    st.markdown(f"**{name}** — {ev['model']} ({ev['accuracy']}) — {ev['notes']}", unsafe_allow_html=True)
            for w in vendors['warnings']: st.warning(w)

        with st.container(border=True):
            st.markdown("<div style='background-color:#F3E5F5;padding:15px;border-radius:10px;border-left:5px solid #6A1B9A'><h2 style='color:#6A1B9A;margin:0'>Layer 4 — TCO & Drift Prediction</h2></div>", unsafe_allow_html=True)
            st.write("")
            # Labels bold + larger than values; values not bold + smaller
            def metric_block(label, value):
                return (f"<p style='font-size:20px;font-weight:800;color:#1B2A4A;margin:0 0 2px 0'>{label}</p>"
                        f"<p style='font-size:15px;font-weight:400;color:#444;margin:0'>{value}</p>")
            t1,t2,t3,t4 = st.columns(4)
            with t1: st.markdown(metric_block("CAPEX Score", f"{tco.capex_score}/30"), unsafe_allow_html=True)
            with t2: st.markdown(metric_block("Calibration", f"Every {tco.calib_months} mo"), unsafe_allow_html=True)
            with t3: st.markdown(metric_block("Liner Life", tco.liner_life), unsafe_allow_html=True)
            with t4: st.markdown(metric_block("Electrode Life", tco.electrode_life), unsafe_allow_html=True)
            if tco.pressure_check: st.info(f"🔒 {tco.pressure_check}")
            st.markdown("<p style='font-size:18px;font-weight:bold'>CAPEX Breakdown</p>", unsafe_allow_html=True)
            bc = st.columns(len(tco.breakdown))
            for i,(k,vs) in enumerate(tco.breakdown.items()):
                with bc[i]: st.markdown(f"**{k}:** {vs}/6 {'🟩'*vs}{'⬜'*(6-vs)}")
            st.markdown("<p style='font-size:18px;font-weight:bold'>Drift Risk Assessment</p>", unsafe_allow_html=True)
            for risk in tco.drift_risks:
                ic = "🔴" if risk.level=="High" else "🟡" if risk.level=="Medium" else "🟢"
                with st.expander(f"{ic} {risk.indicator} — {risk.level}"):
                    st.markdown(f"**Description:** {risk.description}")
                    st.markdown("**Diagnostic Steps:**")
                    for step in risk.steps: st.markdown(f"  {step}")
                    st.markdown("**Recommended Tools:**")
                    for vn,tool in risk.tools.items(): st.markdown(f"  • **{vn}:** {tool}")
                    st.markdown(f"**Frequency:** {risk.frequency}")
            st.markdown("<p style='font-size:18px;font-weight:bold'>📋 Maintenance Plan</p>", unsafe_allow_html=True)
            mnt = tco.maintenance
            for title,items,icon,bg,tc2 in [
                ("Continuous Monitoring",mnt.continuous,"🔄","#E8F5E9","#2E7D32"),
                ("Monthly Checks",mnt.monthly,"📅","#E3F2FD","#1565C0"),
                ("Quarterly Inspection",mnt.quarterly,"🔍","#FFF3E0","#E65100"),
                ("Semi-Annual",mnt.semi_annual,"⚙️","#FCE4D6","#BF360C"),
                ("Annual Maintenance",mnt.annual,"🛠️","#F3E5F5","#6A1B9A"),
                ("Multi-Year (3-5yr)",mnt.multi_year,"📊","#FFF9C4","#F57F17"),
                ("Component Lifespan",mnt.replacement,"♻️","#FFCDD2","#C62828")]:
                if items:
                    st.markdown(f"<div style='background:{bg};padding:12px 15px;border-radius:8px;border-left:4px solid {tc2};margin:8px 0'><p style='font-size:16px;font-weight:bold;color:{tc2};margin:0'>{icon} {title}</p></div>", unsafe_allow_html=True)
                    for item in items: st.markdown(f"&nbsp;&nbsp;&nbsp;&nbsp;• {item}")
            st.markdown("---")

            tag = meta["tag"]

            # ---- Download Excel maintenance sheet ----
            history_for_export = load_history(tag_filter=tag) if tag != "NEW-INSTRUMENT" else []
            checklist_saved = load_checklist_records(tag) if tag != "NEW-INSTRUMENT" else {}
            excel_buf = generate_maintenance_excel(tag, fluid, cat, m, vendors['emerson'], vendors['eh'], vendors['krohne'], tco, history_for_export, checklist_saved)
            st.download_button(label="📄 Download Maintenance Excel", data=excel_buf,
                file_name=f"MagFlow_Maintenance_{tag.replace('/','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            st.caption("💡 The Excel sheet has 3 tabs: Guideline, Maintenance Checklist (month-by-month, "
                       "with a Work done column for monthly & quarterly tasks), and Historical Data.")

            st.markdown("<p style='font-size:18px;font-weight:bold'>🔍 Validation vs JESA Projects</p>", unsafe_allow_html=True)
            if tco.validation:
                for match in tco.validation:
                    ic = "✅" if "MATCH" in match.match_type else ("⚠️" if "PARTIAL" in match.match_type else "❌")
                    with st.expander(f"{ic} {match.match_type} — {match.project} ({match.fluid}) — {match.confidence}%"):
                        for k,vd in match.details.items(): st.markdown(f"**{k}:** {vd}")
                        if match.explanation: st.markdown(f"**Why:** {match.explanation}")
                        if match.cost_impact: st.markdown(f"**Cost:** {match.cost_impact}")
                total = len(tco.validation); matches = sum(1 for mt in tco.validation if mt.confidence>=80)
                st.caption(f"{matches}/{total} matches. {'Validated ✓' if matches>total/2 else 'Review ⚠'}")
            else: st.info("No matching project.")
            imported_matches = compare_imported_projects(fluid, m, meta)
            if imported_matches:
                st.markdown("<p style='font-size:16px;font-weight:bold'>📂 Matches from Imported Projects</p>", unsafe_allow_html=True)
                for im in imported_matches:
                    rec = im["record"]
                    title = f"📂 IMPORTED MATCH — {rec.get('project','Imported project')} | {rec.get('tag','N/A')} ({rec.get('fluid','')}) — {im['confidence']}%"
                    with st.expander(title):
                        st.markdown(f"**Source:** {rec.get('source','Imported Project')}")
                        st.markdown(f"**Tag:** {rec.get('tag','')}")
                        st.markdown(f"**Fluid:** {rec.get('fluid','')}")
                        st.markdown(f"**DN / Flow:** DN{rec.get('dn','')} | {rec.get('flow_normal','')} → {rec.get('flow_max','')} m³/h")
                        st.markdown(f"**Design T/P:** {rec.get('temp_design','')} °C | {rec.get('pressure_design','')} bar")
                        st.markdown(f"**Materials:** {rec.get('electrode','')} / {rec.get('liner','')}")
                        st.markdown("**Why it matched:**")
                        for label, detail in im["details"]:
                            st.markdown(f"- **{label}:** {detail}")
            if tco.notes_response:
                st.markdown("<p style='font-size:18px;font-weight:bold'>💡 AI Response</p>", unsafe_allow_html=True)
                st.markdown(tco.notes_response)
        st.divider()
        st.caption("Source: JESA Internal DB, JESA Flow App 2024, Vendor datasheets")

# ------------------------------------------------------------
#  PROJECT IMPORT
# ------------------------------------------------------------
with mt_import:
    st.header("📂 Project Import — Datasheet Analyzer")
    st.markdown("Upload a JESA flowmeter datasheet PDF. The AI extracts project instruments, recognizes existing fluids/materials even with spelling variants, and sends incomplete new database items to engineering review.")
    col_upload, col_db = st.columns([1, 1])
    with col_upload:
        st.subheader("📤 Upload Datasheet PDF")
        project_name_override = st.text_input("Project Name (optional)", placeholder="e.g. Central Axis Program / SAFI")
        uploaded_pdf = st.file_uploader("Upload JESA Datasheet PDF", type=["pdf"], key="pdf_import")
        if uploaded_pdf:
            st.success(f"✅ **{uploaded_pdf.name}** ready")
            if st.button("🤖 Analyze & Check Database", type="primary", use_container_width=True):
                with st.spinner("Step 1/3 — AI reading datasheet..."):
                    pdf_bytes = uploaded_pdf.read()
                    instruments, error = extract_datasheet_with_ai(pdf_bytes)
                if error or not instruments:
                    st.error(f"❌ Extraction failed: {error or 'No instruments found'}")
                else:
                    with st.spinner("Step 2/3 — Loading current database from GitHub..."):
                        wb, sha, err = get_excel_from_github()
                    if err:
                        st.error(f"❌ Cannot load database: {err}")
                    else:
                        with st.spinner("Step 3/3 — Matching names and detecting review items..."):
                            existing = get_existing_materials(wb)
                            instruments, matched_notes = normalize_imported_instruments(instruments, existing)
                            new_findings = detect_new_materials(instruments, existing)
                        st.session_state['import_instruments'] = instruments
                        st.session_state['import_wb'] = wb
                        st.session_state['import_sha'] = sha
                        st.session_state['import_new'] = new_findings
                        st.session_state['import_matches'] = matched_notes
                        st.session_state['import_project'] = project_name_override or (instruments[0].get('project','') if instruments else '')
                        st.success(f"✅ Found **{len(instruments)}** instrument(s) — **{len(new_findings)}** item(s) need review")
        if 'import_instruments' in st.session_state:
            instruments = st.session_state['import_instruments']
            new_findings = st.session_state['import_new']
            matched_notes = st.session_state.get('import_matches', [])
            pname = st.session_state['import_project']
            st.divider()
            st.subheader("📋 Extracted Instruments")
            if matched_notes:
                with st.expander("✅ Existing database matches / aliases recognized", expanded=False):
                    for note in matched_notes:
                        st.markdown(f"- {note}")
            for inst in instruments:
                with st.expander(f"🔧 {inst.get('tag','N/A')} — {inst.get('fluid','N/A')}"):
                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown(f"**Project:** {inst.get('project', pname)}"); st.markdown(f"**Fluid:** {inst.get('fluid','')}")
                        if inst.get('fluid_original') and inst.get('fluid_original') != inst.get('fluid'):
                            st.caption(f"PDF value: {inst.get('fluid_original')} → database value: {inst.get('fluid')}")
                        st.markdown(f"**DN:** {inst.get('dn','')} mm"); st.markdown(f"**Flow:** {inst.get('flow_normal','')} → {inst.get('flow_max','')} m³/h")
                        st.markdown(f"**Temp design:** {inst.get('temp_design','')} °C"); st.markdown(f"**Pressure design:** {inst.get('pressure_design','')} bar-g")
                    with c2:
                        st.markdown(f"**Electrode:** {inst.get('electrode','')}"); st.markdown(f"**Liner:** {inst.get('liner','')}")
                        st.markdown(f"**Tube:** {inst.get('tube','')}"); st.markdown(f"**Grounding:** {inst.get('grounding','')}")
                        st.markdown(f"**Accuracy:** {inst.get('accuracy','')}"); st.markdown(f"**Vendor/Model:** {inst.get('vendor','')} / {inst.get('model','')}")
            if new_findings:
                st.divider()
                st.subheader("🧾 Database Review Items")
                st.warning(f"**{len(new_findings)} item(s)** need I&C review before entering the curated database:")
                approved = []
                for i, finding in enumerate(new_findings):
                    icons = {"electrode": "⚡", "liner": "🟡", "fluid": "💧", "fluid_electrode_variant": "🔀", "fluid_liner_variant": "🔀"}
                    labels = {"electrode": "New Electrode", "liner": "New Liner", "fluid": "New Fluid", "fluid_electrode_variant": "Electrode Variant", "fluid_liner_variant": "Liner Variant"}
                    checked = st.checkbox(f"{icons.get(finding['type'],'🆕')} **{labels.get(finding['type'],'New')}:** `{finding['value']}` — {finding['context']}", value=True, key=f"approve_{i}")
                    st.caption(finding.get("reason", "Needs review before database insertion."))
                    if checked:
                        approved.append(finding)
                st.session_state['import_approved'] = approved
            else:
                st.info("✅ No new database review items detected. Existing spelling variants were matched to the current database.")
                st.session_state['import_approved'] = []
            st.divider()
            col_save1, col_save2 = st.columns(2)
            with col_save1:
                if st.button("💾 Save to Project Database", use_container_width=True):
                    with st.spinner("Saving to Google Sheets..."):
                        saved, err = save_project_to_gsheet(instruments, pname)
                    if err: st.error(f"❌ {err}")
                    else: st.success(f"✅ {saved} instrument(s) saved!")
            with col_save2:
                approved_updates = st.session_state.get('import_approved', [])
                if approved_updates:
                    st.info(f"💡 **{len(approved_updates)} selected item(s)** will be saved to `AI Import Review Queue`, not directly into the curated material tables. This avoids adding incomplete records when the PDF does not contain maintenance plan, CAPEX score, or full compatibility rules.")
                    if st.button("📌 Send Selected Items to Review Queue", use_container_width=True):
                        with st.spinner("Updating Data_Collection_v2.xlsx review queue on GitHub..."):
                            wb = st.session_state.get('import_wb')
                            sha = st.session_state.get('import_sha')
                            wb, changes = apply_updates_to_excel(wb, approved_updates, pname)
                            ok, err = push_excel_to_github(wb, sha, f"Add AI import review items for {pname or 'project'}")
                        if err:
                            st.error(f"❌ {err}")
                        else:
                            st.success(f"✅ {len(changes)} item(s) queued for I&C review.")
                            for change in changes:
                                st.markdown(f"- {change}")
                else:
                    st.success("✅ All materials already in database.")
    with col_db:
        st.subheader("📊 Imported Projects Database")
        with st.spinner("Loading..."):
            project_records = load_project_data_from_gsheet()
        if project_records:
            st.caption(f"Total instruments imported: **{len(project_records)}**")
            projects = list(set(r.get('Project','') for r in project_records if r.get('Project','')))
            for proj in sorted(projects):
                proj_insts = [r for r in project_records if r.get('Project','') == proj]
                with st.expander(f"📁 {proj} — {len(proj_insts)} instrument(s)"):
                    for r in proj_insts:
                        st.markdown(f"**{r.get('Tag','')}** | {r.get('Fluid','')} | DN{r.get('DN','')} | ⚡{r.get('Electrode','')} / 🟡{r.get('Liner','')} | _{r.get('Import Date','')}_")
        else: st.info("No projects imported yet. Upload a datasheet to get started.")
